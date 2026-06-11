"""
Export routes — JSON, CSV-zip, cohort XLSX, diagnoses list, drugs list.
Extracted from server.py for maintainability.
"""
import io
import re
import csv
import json
import zipfile
import logging
from datetime import datetime, timezone
from typing import List, Optional, Dict, Any

from fastapi import APIRouter, Depends, Response
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from database import db
from auth_utils import get_current_user

logger = logging.getLogger(__name__)

router = APIRouter()

COLLECTIONS_TO_EXPORT = [
    "patients", "assessments", "criteria_evaluations",
    "therapies", "lab_exams", "reminders", "sclero_profiles",
    "disease_profiles",
]


async def _collect_org_data(organization_id: str) -> dict:
    data = {}
    for col in COLLECTIONS_TO_EXPORT:
        docs = await db[col].find({"organization_id": organization_id}, {"_id": 0}).to_list(100000)
        data[col] = docs
    return data


@router.get("/export/json")
async def export_json(user: dict = Depends(get_current_user)):
    """Export all org data as a single JSON file."""
    data = await _collect_org_data(user["organization_id"])
    payload = {
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "exported_by": user.get("name"),
        "organization_id": user["organization_id"],
        "version": 1,
        **data,
    }
    body = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
    fname = f"clinimetria_export_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.json"
    return Response(
        content=body,
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


def _flatten_for_csv(value):
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value if value is not None else ""


@router.get("/export/csv-zip")
async def export_csv_zip(user: dict = Depends(get_current_user)):
    """Export all org data as a ZIP of CSV files (one per collection)."""
    data = await _collect_org_data(user["organization_id"])
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for col, rows in data.items():
            csv_buf = io.StringIO()
            if rows:
                keys = []
                seen = set()
                for r in rows:
                    for k in r.keys():
                        if k not in seen:
                            seen.add(k)
                            keys.append(k)
                writer = csv.DictWriter(csv_buf, fieldnames=keys, extrasaction="ignore")
                writer.writeheader()
                for r in rows:
                    writer.writerow({k: _flatten_for_csv(r.get(k)) for k in keys})
            else:
                csv_buf.write("(no records)\n")
            zf.writestr(f"{col}.csv", csv_buf.getvalue())
        manifest = {
            "exported_at": datetime.now(timezone.utc).isoformat(),
            "exported_by": user.get("name"),
            "organization_id": user["organization_id"],
            "counts": {col: len(rows) for col, rows in data.items()},
        }
        zf.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))

    buf.seek(0)
    fname = f"clinimetria_export_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.zip"
    return Response(
        content=buf.read(),
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ── Cohort XLSX helpers ───────────────────────────────────────────────────────
# _episode_state_at is the canonical point-in-time reconstruction function,
# defined once in helpers.py and imported here.
from helpers import _episode_state_at


def _therapies_active_on(therapies: List[dict], date_str: str) -> List[dict]:
    """
    Return therapies active at date_str with historically-accurate dose/frequency/route.
    Delegates to _episode_state_at() (helpers.py) for point-in-time reconstruction.
    """
    out = []
    for t in therapies:
        state = _episode_state_at(t, date_str)
        if state is not None:
            out.append(state)
    return out


def _format_therapy(t: dict) -> str:
    parts = [t.get("drug_name") or "?"]
    if t.get("dose"):
        parts.append(t["dose"])
    if t.get("frequency"):
        parts.append(t["frequency"])
    if t.get("route"):
        parts.append(t["route"])
    return " ".join(parts)


def _days_delta(anchor_iso: Optional[str], visit_iso: Optional[str]):
    if not anchor_iso or not visit_iso:
        return ""
    try:
        a = datetime.fromisoformat(anchor_iso[:10]).date()
        v = datetime.fromisoformat(visit_iso[:10]).date()
        return (v - a).days
    except Exception:
        return ""


async def _load_cohort_data(org_id: str, diagnosis: Optional[str]) -> tuple:
    patient_query: dict = {"organization_id": org_id}
    if diagnosis and diagnosis.strip():
        patient_query["diagnosi"] = {"$regex": re.escape(diagnosis.strip()), "$options": "i"}
    patients = await db.patients.find(patient_query, {"_id": 0}).to_list(100000)
    patient_ids = [p["id"] for p in patients]

    assess_by_pid: Dict[str, List[dict]] = {}
    therap_by_pid: Dict[str, List[dict]] = {}
    prof_by_pid: Dict[str, Dict[str, dict]] = {}
    sclero_by_pid: Dict[str, dict] = {}

    if patient_ids:
        flt = {"organization_id": org_id, "patient_id": {"$in": patient_ids}}
        for a in await db.assessments.find(flt, {"_id": 0}).to_list(1000000):
            assess_by_pid.setdefault(a["patient_id"], []).append(a)
        for t in await db.therapies.find(flt, {"_id": 0}).to_list(1000000):
            therap_by_pid.setdefault(t["patient_id"], []).append(t)
        for d in await db.disease_profiles.find(flt, {"_id": 0}).to_list(1000000):
            prof_by_pid.setdefault(d["patient_id"], {})[d.get("disease_type", "unknown")] = d
        for s in await db.sclero_profiles.find(flt, {"_id": 0}).to_list(1000000):
            sclero_by_pid[s["patient_id"]] = s

    return patients, assess_by_pid, therap_by_pid, prof_by_pid, sclero_by_pid


def _build_visits_pivot(assess_by_pid: Dict[str, List[dict]]) -> tuple:
    visits_by_pid: Dict[str, List[dict]] = {}
    max_visits = 0
    for pid, alist in assess_by_pid.items():
        by_date: Dict[str, List[dict]] = {}
        for a in alist:
            d = (a.get("date") or a.get("created_at") or "")[:10]
            if d:
                by_date.setdefault(d, []).append(a)
        visits = sorted(({"date": d, "assessments": ass} for d, ass in by_date.items()), key=lambda v: v["date"])
        visits_by_pid[pid] = visits
        if len(visits) > max_visits:
            max_visits = len(visits)

    seen_idx: set = set()
    all_indices: List[str] = []
    for alist in assess_by_pid.values():
        for a in alist:
            it = a.get("index_type")
            if it and it not in seen_idx:
                seen_idx.add(it)
                all_indices.append(it)
    all_indices.sort()
    return visits_by_pid, max_visits, all_indices


def _compute_anchor_dates(therap_by_pid: Dict[str, List[dict]], anchor_drug: str) -> Dict[str, str]:
    anchor_norm = (anchor_drug or "").strip().lower()
    out: Dict[str, str] = {}
    if not anchor_norm:
        return out
    for pid, ther_list in therap_by_pid.items():
        matches = [
            t for t in ther_list
            if t.get("start_date") and anchor_norm in (t.get("drug_name") or "").lower()
        ]
        if matches:
            out[pid] = min(t["start_date"] for t in matches)
    return out


def _collect_profile_columns(
    prof_by_pid: Dict[str, Dict[str, dict]],
    sclero_by_pid: Dict[str, dict],
) -> tuple:
    profile_cols: List[tuple] = []
    profile_keys_seen: set = set()
    for prof_map in prof_by_pid.values():
        for dtype, doc in prof_map.items():
            for k in (doc.get("data") or {}).keys():
                tag = (dtype, k)
                if tag not in profile_keys_seen:
                    profile_keys_seen.add(tag)
                    profile_cols.append(tag)

    sclero_cols: List[tuple] = []
    sclero_keys_seen: set = set()
    for sdoc in sclero_by_pid.values():
        for sec in ("cutaneous", "antibody", "vascular", "ild", "pah", "gi", "msk"):
            section_data = sdoc.get(sec) or {}
            if not isinstance(section_data, dict):
                continue
            for k in section_data.keys():
                tag = (sec, k)
                if tag not in sclero_keys_seen:
                    sclero_keys_seen.add(tag)
                    sclero_cols.append(tag)
    return profile_cols, sclero_cols


def _build_cohort_columns(
    pseudo: bool,
    profile_cols: List[tuple],
    sclero_cols: List[tuple],
    anchor_norm: bool,
    max_visits: int,
    all_indices: List[str],
) -> List[str]:
    cols: List[str] = ["codice_paziente"]
    if not pseudo:
        cols.extend(["cognome", "nome", "codice_fiscale", "data_nascita"])
    cols.extend(["anno_nascita", "sesso", "diagnosi", "note", "n_visite"])
    for dtype, k in profile_cols:
        cols.append(f"profilo_{dtype}__{k}")
    for sec, k in sclero_cols:
        cols.append(f"ssc_{sec}__{k}")
    if anchor_norm:
        cols.extend(["anchor_drug", "anchor_t0"])
    for i in range(1, max_visits + 1):
        cols.append(f"t{i}_data")
        if anchor_norm:
            cols.append(f"t{i}_giorni_da_anchor")
        for idx in all_indices:
            cols.append(f"t{i}_{idx}_score")
        cols.append(f"t{i}_terapie_attive")
    return cols


def _cell(val):
    if val is None:
        return ""
    if isinstance(val, bool):
        return "sì" if val else "no"
    if isinstance(val, (dict, list)):
        return json.dumps(val, ensure_ascii=False)
    return val


def _build_patient_row(
    p: dict,
    *,
    pseudo: bool,
    visits: List[dict],
    all_ther: List[dict],
    profile_cols: List[tuple],
    sclero_cols: List[tuple],
    prof_map: Dict[str, dict],
    sdoc: Optional[dict],
    all_indices: List[str],
    max_visits: int,
    anchor_drug: Optional[str],
    anchor_t0: Optional[str],
) -> List[Any]:
    row: List[Any] = [p.get("codice_paziente") or ""]
    if not pseudo:
        row.extend([
            p.get("cognome") or "",
            p.get("nome") or "",
            p.get("codice_fiscale") or "",
            p.get("data_nascita") or "",
        ])
    row.extend([
        p.get("anno_nascita") or "",
        p.get("sesso") or "",
        p.get("diagnosi") or "",
        (p.get("note") or "").replace("\n", " "),
        len(visits),
    ])
    for dtype, k in profile_cols:
        doc = prof_map.get(dtype)
        val = (doc or {}).get("data", {}).get(k) if doc else None
        row.append(_cell(val))
    for sec, k in sclero_cols:
        val = None
        if sdoc:
            section_data = sdoc.get(sec) or {}
            if isinstance(section_data, dict):
                val = section_data.get(k)
        row.append(_cell(val))
    if anchor_drug:
        row.append(anchor_drug)
        row.append(anchor_t0 or "")
    for i in range(max_visits):
        if i < len(visits):
            v = visits[i]
            row.append(v["date"])
            if anchor_drug:
                row.append(_days_delta(anchor_t0, v["date"]))
            score_map: Dict[str, Any] = {}
            for a in v["assessments"]:
                it = a.get("index_type")
                s = a.get("score")
                if it and s is not None and it not in score_map:
                    score_map[it] = s
            for idx in all_indices:
                row.append(score_map.get(idx, ""))
            active = _therapies_active_on(all_ther, v["date"])
            row.append("; ".join(_format_therapy(t) for t in active) if active else "")
        else:
            row.append("")
            if anchor_drug:
                row.append("")
            for _ in all_indices:
                row.append("")
            row.append("")
    return row


def _style_header_row(ws, ncols: int, row: int = 1) -> None:
    hfont = Font(bold=True, color="FFFFFF", size=10)
    hfill = PatternFill("solid", fgColor="0A2540")
    halign = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for col_idx in range(1, ncols + 1):
        c = ws.cell(row=row, column=col_idx)
        c.font = hfont
        c.fill = hfill
        c.alignment = halign
    ws.row_dimensions[row].height = 30
    ws.freeze_panes = "A2"


def _autosize_columns(ws, cols: List[str]) -> None:
    for col_idx, name in enumerate(cols, start=1):
        max_len = len(name)
        letter = ws.cell(row=1, column=col_idx).column_letter
        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=col_idx).value
            if val is not None:
                L = len(str(val))
                if L > max_len:
                    max_len = L
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 60)


def _write_long_sheet(
    wb,
    *,
    title: str,
    columns: List[str],
    pseudo: bool,
    patients: List[dict],
    items_for_patient,
    sort_key=None,
) -> None:
    cols = [c for c in columns if not (pseudo and c in ("cognome", "nome"))]
    ws = wb.create_sheet(title)
    ws.append(cols)
    _style_header_row(ws, len(cols))
    ws.row_dimensions[1].height = 25
    for p in patients:
        items = items_for_patient(p)
        if sort_key:
            items = sorted(items, key=sort_key)
        for it in items:
            row = []
            for k in cols:
                if k == "codice_paziente":
                    row.append(p.get("codice_paziente") or "")
                elif k in ("cognome", "nome", "diagnosi"):
                    row.append(p.get(k) or "")
                else:
                    row.append(_cell(it.get(k)))
            ws.append(row)


@router.get("/export/cohort-xlsx")
async def export_cohort_xlsx(
    diagnosis: Optional[str] = None,
    anchor_drug: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    """
    Export an Excel workbook with one row per patient (optionally filtered by
    diagnosis), demographics + disease profile + pivoted visits showing scores
    and active therapies. If anchor_drug is set, day-deltas are relative to
    the earliest start of that drug.
    """
    org_id = user["organization_id"]

    patients, assess_by_pid, therap_by_pid, prof_by_pid, sclero_by_pid = await _load_cohort_data(
        org_id, diagnosis
    )
    visits_by_pid, max_visits, all_indices = _build_visits_pivot(assess_by_pid)
    anchor_by_pid = _compute_anchor_dates(therap_by_pid, anchor_drug or "")
    profile_cols, sclero_cols = _collect_profile_columns(prof_by_pid, sclero_by_pid)

    org = await db.organizations.find_one({"id": org_id}, {"_id": 0})
    pseudo = bool((org or {}).get("pseudonymized_mode"))
    anchor_active = bool((anchor_drug or "").strip())

    wb = Workbook()
    ws = wb.active
    ws.title = "Coorte"

    cols = _build_cohort_columns(pseudo, profile_cols, sclero_cols, anchor_active, max_visits, all_indices)
    ws.append(cols)
    _style_header_row(ws, len(cols))

    for p in patients:
        ws.append(_build_patient_row(
            p,
            pseudo=pseudo,
            visits=visits_by_pid.get(p["id"], []),
            all_ther=therap_by_pid.get(p["id"], []),
            profile_cols=profile_cols,
            sclero_cols=sclero_cols,
            prof_map=prof_by_pid.get(p["id"], {}),
            sdoc=sclero_by_pid.get(p["id"]),
            all_indices=all_indices,
            max_visits=max_visits,
            anchor_drug=(anchor_drug if anchor_active else None),
            anchor_t0=anchor_by_pid.get(p["id"]),
        ))
    _autosize_columns(ws, cols)

    _write_long_sheet(
        wb,
        title="Terapie",
        columns=[
            "codice_paziente", "cognome", "nome", "diagnosi", "drug_name", "category",
            "indication", "dose", "frequency", "route", "start_date", "end_date",
            "status", "discontinuation_reason", "auto_discontinued", "notes", "created_by_name",
        ],
        pseudo=pseudo,
        patients=patients,
        items_for_patient=lambda p: therap_by_pid.get(p["id"], []),
    )
    _write_long_sheet(
        wb,
        title="Valutazioni",
        columns=[
            "codice_paziente", "cognome", "nome", "diagnosi", "date",
            "index_type", "score", "interpretation",
            "tender_count", "swollen_count", "created_by_name", "notes",
        ],
        pseudo=pseudo,
        patients=patients,
        items_for_patient=lambda p: assess_by_pid.get(p["id"], []),
        sort_key=lambda x: (x.get("date") or ""),
    )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    diag_tag = (diagnosis or "all").strip().replace(" ", "_").replace("/", "_")[:30] or "all"
    fname = f"coorte_{diag_tag}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/export/diagnoses")
async def export_diagnoses(user: dict = Depends(get_current_user)):
    """Return distinct diagnoses for the organization (cohort filter)."""
    org_id = user["organization_id"]
    vals = await db.patients.distinct("diagnosi", {"organization_id": org_id})
    clean = sorted({(v or "").strip() for v in vals if (v or "").strip()})
    return {"diagnoses": clean}


@router.get("/export/drugs")
async def export_drugs(user: dict = Depends(get_current_user)):
    """Return distinct drug names used in therapies (cohort anchor selection)."""
    org_id = user["organization_id"]
    vals = await db.therapies.distinct("drug_name", {"organization_id": org_id})
    clean = sorted({(v or "").strip() for v in vals if (v or "").strip()}, key=lambda s: s.lower())
    return {"drugs": clean}
