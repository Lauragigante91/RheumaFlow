"""Iteration 6 regression tests for the post-refactor changes:

1. GET /api/auth/me returns 'is_demo' boolean (false for admin).
2. POST /api/auth/demo creates a fresh demo user, and GET /api/auth/me on
   that session returns 'is_demo': true.
3. GET /api/export/cohort-xlsx (no filters) returns a 200 + valid xlsx with
   sheets ['Coorte','Terapie','Valutazioni']; required headers exist on Coorte.
4. GET /api/export/cohort-xlsx?anchor_drug=Methotrexate returns 200 with
   anchor_drug, anchor_t0, t1_giorni_da_anchor columns; deltas are int or "".
5. GET /api/export/cohort-xlsx?diagnosis=Artrite Reumatoide only includes
   patients whose diagnosi contains the filter (case-insensitive).
"""

import io
import os
import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")
if not BASE_URL:
    with open("/app/frontend/.env") as f:
        for line in f:
            if line.startswith("REACT_APP_BACKEND_URL="):
                BASE_URL = line.split("=", 1)[1].strip().rstrip("/")
API = f"{BASE_URL}/api"

ADMIN_EMAIL = "admin@clinimetria.it"
ADMIN_PASSWORD = "admin123"


# ---------- fixtures ----------

@pytest.fixture(scope="module")
def admin_session():
    s = requests.Session()
    s.headers.update({"Content-Type": "application/json"})
    r = s.post(f"{API}/auth/login", json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD})
    assert r.status_code == 200, f"admin login failed: {r.status_code} {r.text}"
    yield s


# ---------- auth/me + is_demo ----------

class TestAuthMeIsDemo:
    def test_admin_me_has_is_demo_false(self, admin_session):
        r = admin_session.get(f"{API}/auth/me")
        assert r.status_code == 200, r.text
        data = r.json()
        assert "is_demo" in data, "auth/me missing is_demo field"
        assert data["is_demo"] is False
        # Sanity: still has the standard fields
        assert data["email"].lower() == ADMIN_EMAIL
        assert "organization_id" in data

    def test_demo_session_me_has_is_demo_true(self):
        # POST /api/auth/demo creates a fresh demo session via cookies
        s = requests.Session()
        s.headers.update({"Content-Type": "application/json"})
        r = s.post(f"{API}/auth/demo")
        assert r.status_code == 200, f"demo login failed: {r.status_code} {r.text}"
        body = r.json()
        assert body.get("is_demo") is True

        # Now /auth/me with the same cookie jar should mirror is_demo true
        r2 = s.get(f"{API}/auth/me")
        assert r2.status_code == 200, r2.text
        data = r2.json()
        assert data.get("is_demo") is True
        assert data["email"].endswith("@clinimetria.demo")
        # Cleanup: best-effort logout (org persists, but session token is invalidated)
        try:
            s.post(f"{API}/auth/logout")
        except Exception:
            pass


# ---------- export_cohort_xlsx (refactored into 8 helpers) ----------

REQUIRED_COORTE_HEADERS = {
    "codice_paziente", "anno_nascita", "sesso", "diagnosi",
    "t1_data", "t1_terapie_attive",
}


def _download_xlsx(session, params=None):
    r = session.get(f"{API}/export/cohort-xlsx", params=params or {})
    assert r.status_code == 200, f"export failed: {r.status_code} {r.text[:300]}"
    assert r.headers.get("content-type", "").startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ), f"unexpected content-type: {r.headers.get('content-type')}"
    wb = load_workbook(io.BytesIO(r.content))
    return wb


def _headers_of(ws):
    return [c.value for c in ws[1] if c.value is not None]


class TestExportCohortRefactor:
    def test_export_no_filter_structure(self, admin_session):
        wb = _download_xlsx(admin_session)
        # Sheets present
        assert set(wb.sheetnames) >= {"Coorte", "Terapie", "Valutazioni"}, wb.sheetnames

        ws = wb["Coorte"]
        headers = _headers_of(ws)
        missing = REQUIRED_COORTE_HEADERS - set(headers)
        assert not missing, f"Coorte sheet missing required headers: {missing}; got={headers}"

        # at least the header row exists; data rows may be 0 if no patients
        assert ws.max_row >= 1

    def test_export_anchor_drug_methotrexate(self, admin_session):
        wb = _download_xlsx(admin_session, params={"anchor_drug": "Methotrexate"})
        ws = wb["Coorte"]
        headers = _headers_of(ws)

        # Must contain anchor metadata columns
        assert "anchor_drug" in headers, headers
        assert "anchor_t0" in headers, headers
        # And at least one tN_giorni_da_anchor column
        delta_cols = [h for h in headers if isinstance(h, str) and h.endswith("_giorni_da_anchor")]
        assert delta_cols, f"no t*_giorni_da_anchor columns; got={headers}"

        # Validate column data type for the first delta col across all rows:
        # must be int or "" (empty / None) — never garbage strings.
        col_idx = headers.index(delta_cols[0]) + 1
        for row in range(2, ws.max_row + 1):
            v = ws.cell(row=row, column=col_idx).value
            assert v is None or v == "" or isinstance(v, int), (
                f"row {row} {delta_cols[0]} has non-int delta: {v!r}"
            )

    def test_export_diagnosis_filter_artrite_reumatoide(self, admin_session):
        wb = _download_xlsx(admin_session, params={"diagnosis": "Artrite Reumatoide"})
        ws = wb["Coorte"]
        headers = _headers_of(ws)
        assert "diagnosi" in headers
        diag_idx = headers.index("diagnosi") + 1

        # Every patient row must have a diagnosi containing 'artrite reumatoide' (ci)
        non_matching = []
        for row in range(2, ws.max_row + 1):
            d = ws.cell(row=row, column=diag_idx).value or ""
            if "artrite reumatoide" not in str(d).lower():
                non_matching.append((row, d))
        assert not non_matching, f"diagnosis filter leaked rows: {non_matching[:5]}"

    def test_export_pseudonymized_mode_skips_pii(self, admin_session):
        """If admin org is in pseudonymized_mode, cognome/nome must be absent on
        all 3 sheets. Otherwise we just skip (no PSEUDO org available)."""
        # discover current admin org
        me = admin_session.get(f"{API}/auth/me").json()
        if not me.get("pseudonymized_mode"):
            pytest.skip("admin org is not in pseudonymized_mode; skip PII-masking check")

        wb = _download_xlsx(admin_session)
        for sheet_name in ("Coorte", "Terapie", "Valutazioni"):
            ws = wb[sheet_name]
            headers = _headers_of(ws)
            assert "cognome" not in headers, f"{sheet_name} leaked cognome"
            assert "nome" not in headers, f"{sheet_name} leaked nome"
