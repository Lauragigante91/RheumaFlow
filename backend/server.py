from dotenv import load_dotenv
from pathlib import Path
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

import os
import logging
import secrets
import uuid
from datetime import datetime, timezone, timedelta
from typing import List, Optional, Dict, Any

import re
import bcrypt
import jwt
import json
import csv
import io
import zipfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import httpx
from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, Response, UploadFile, File
from fastapi.responses import StreamingResponse, JSONResponse, FileResponse
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, ConfigDict, EmailStr

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

JWT_ALGORITHM = "HS256"


def get_jwt_secret() -> str:
    return os.environ["JWT_SECRET"]


# ==================== AUTH HELPERS ====================
def hash_password(password: str) -> str:
    salt = bcrypt.gensalt()
    return bcrypt.hashpw(password.encode("utf-8"), salt).decode("utf-8")


def verify_password(plain: str, hashed: str) -> bool:
    return bcrypt.checkpw(plain.encode("utf-8"), hashed.encode("utf-8"))


def create_access_token(user_id: str, email: str, organization_id: str) -> str:
    payload = {
        "sub": user_id,
        "email": email,
        "org": organization_id,
        "exp": datetime.now(timezone.utc) + timedelta(hours=8),
        "type": "access",
    }
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)


def create_refresh_token(user_id: str) -> str:
    payload = {
        "sub": user_id,
        "exp": datetime.now(timezone.utc) + timedelta(days=7),
        "type": "refresh",
    }
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)


async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Non autenticato")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Tipo di token non valido")
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "password_hash": 0})
        if not user:
            raise HTTPException(status_code=401, detail="Utente non trovato")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token scaduto")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token non valido")


def set_auth_cookies(response: Response, access_token: str, refresh_token: str):
    # Cancella eventuali cookie obsoleti (es. con SameSite/Secure differenti) prima di settare i nuovi
    response.delete_cookie(key="access_token", path="/")
    response.delete_cookie(key="refresh_token", path="/")
    response.set_cookie(key="access_token", value=access_token, httponly=True, secure=True, samesite="none", max_age=28800, path="/")
    response.set_cookie(key="refresh_token", value=refresh_token, httponly=True, secure=True, samesite="none", max_age=604800, path="/")


# ==================== MODELS ====================
class Organization(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    invite_code: str = Field(default_factory=lambda: secrets.token_urlsafe(8))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    pseudonymized_mode: bool = False  # se True l'UI nasconde nome/cognome/CF/data_nascita
    ai_strict_privacy: bool = False   # se True de-identifica il testo prima dell'AI e non archivia documenti sorgente


class OrganizationSettings(BaseModel):
    pseudonymized_mode: Optional[bool] = None
    ai_strict_privacy: Optional[bool] = None


class UserPublic(BaseModel):
    id: str
    email: str
    name: str
    role: str
    organization_id: str
    organization_name: Optional[str] = None


class RegisterRequest(BaseModel):
    email: str
    password: str
    name: str
    organization_name: Optional[str] = None  # se nuovo
    invite_code: Optional[str] = None  # se join


class LoginRequest(BaseModel):
    email: str
    password: str


class PatientBase(BaseModel):
    # Pseudonimizzazione: il codice_paziente è sempre preferito.
    # Se modalità pseudonimizzata attiva nell'org, i campi nominativi NON devono essere inviati dall'UI.
    codice_paziente: Optional[str] = None  # identificatore scelto dall'utente (es. RX-2026-001)
    nome: Optional[str] = None
    cognome: Optional[str] = None
    anno_nascita: Optional[int] = None  # alternativa meno identificante a data_nascita
    data_nascita: Optional[str] = None
    sesso: Optional[str] = None
    codice_fiscale: Optional[str] = None
    diagnosi: Optional[str] = None
    diagnosi_secondarie: List[str] = Field(default_factory=list)  # overlap diagnoses (es. fibromialgia, osteoporosi)
    note: Optional[str] = None
    # Disease onset — strongly influences classification, prognosis and treatment strategy
    onset_year: Optional[int] = None   # anno esordio di malattia (sempre)
    onset_month: Optional[int] = None  # mese esordio (1-12, facoltativo)
    # Clinical workflow state machine
    # None              → prima visita non ancora effettuata
    # "workup_in_progress" → prima visita effettuata, iter diagnostico in corso
    # "follow_up"          → diagnosi definitiva stabilita, follow-up malattia-specifica
    patient_state: Optional[str] = None


class Patient(PatientBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    organization_id: str
    created_by: str
    created_by_name: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    # Recall flag for "Pazienti da ricontrollare" widget. Stored as nested dict
    # {flag, note, set_at, set_by, set_by_name} where flag ∈ {'private', 'shared'}.
    recall: Optional[Dict[str, Any]] = None


class PatientUpdate(BaseModel):
    codice_paziente: Optional[str] = None
    nome: Optional[str] = None
    cognome: Optional[str] = None
    anno_nascita: Optional[int] = None
    data_nascita: Optional[str] = None
    sesso: Optional[str] = None
    codice_fiscale: Optional[str] = None
    diagnosi: Optional[str] = None
    diagnosi_secondarie: Optional[List[str]] = None
    note: Optional[str] = None
    onset_year: Optional[int] = None
    onset_month: Optional[int] = None
    patient_state: Optional[str] = None


class AssessmentBase(BaseModel):
    patient_id: str
    index_type: str
    date: str
    inputs: Dict[str, Any] = {}
    score: Optional[float] = None
    interpretation: Optional[str] = None
    tender_joints: List[str] = []
    swollen_joints: List[str] = []
    notes: Optional[str] = None
    visit_id: Optional[str] = None    # links this assessment to a specific visit record
    visit_type: Optional[str] = None  # "workup" | "followup" | "prima_visita"


class Assessment(AssessmentBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    organization_id: str
    created_by: str
    created_by_name: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class CriteriaEvaluationBase(BaseModel):
    patient_id: str
    criteria_id: str
    criteria_name: str
    source: str
    date: str
    score: float
    threshold: float
    meets: bool
    selections: Dict[str, Any] = {}
    notes: Optional[str] = None


class CriteriaEvaluation(CriteriaEvaluationBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    organization_id: str
    created_by: str
    created_by_name: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class TherapyEvent(BaseModel):
    """A single clinical event inside a therapy episode."""
    type: str
    date: str
    voided: Optional[bool] = None
    dose: Optional[str] = None
    dose_before: Optional[str] = None
    dose_after: Optional[str] = None
    frequency_before: Optional[str] = None
    frequency_after: Optional[str] = None
    route_before: Optional[str] = None
    route_after: Optional[str] = None
    status_before: Optional[str] = None
    status_after: Optional[str] = None
    reason: Optional[str] = None
    created_by: Optional[str] = None
    created_by_name: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    notes: Optional[str] = None


class TherapyBase(BaseModel):
    patient_id: str
    drug_name: str
    category: str  # csDMARD, bDMARD, tsDMARD, glucocorticoid, NSAID, analgesic, supportive, other
    dose: Optional[str] = None
    frequency: Optional[str] = None
    route: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    status: str = "active"  # active, discontinued, completed
    discontinuation_reason: Optional[str] = None
    auto_discontinued: Optional[bool] = None
    notes: Optional[str] = None
    events: List[TherapyEvent] = Field(default_factory=list)


class Therapy(TherapyBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    organization_id: str
    created_by: str
    created_by_name: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class TherapyUpdate(BaseModel):
    drug_name: Optional[str] = None
    category: Optional[str] = None
    dose: Optional[str] = None
    frequency: Optional[str] = None
    route: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    status: Optional[str] = None
    discontinuation_reason: Optional[str] = None
    notes: Optional[str] = None


class LabExamBase(BaseModel):
    patient_id: str
    date: str
    panel: str  # autoanticorpi, complemento, fase_acuta, emocromo, funzione, urine, custom
    values: Dict[str, Any] = {}  # { test_key: { value, unit, status (positive/negative/normal/high/low), notes? } }
    notes: Optional[str] = None


class LabExam(LabExamBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    organization_id: str
    created_by: str
    created_by_name: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class SpecialistVisitBase(BaseModel):
    patient_id: str
    visit_date: str
    visit_type: str                         # es. "Visita pneumologica", "Consulenza dermatologica"
    specialty: Optional[str] = None         # chiave normalizzata: pneumologia, cardiologia, …
    source_text: Optional[str] = None       # testo completo del referto / consulenza
    sintesi: Optional[str] = None           # sintesi clinica a una riga


class SpecialistVisit(SpecialistVisitBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    organization_id: str
    created_by: str
    created_by_name: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class ReminderBase(BaseModel):
    patient_id: Optional[str] = None  # null = "free" task non legato a un paziente
    due_date: str
    title: str
    type: Optional[str] = None  # follow_up, lab, imaging, therapy, other
    notes: Optional[str] = None
    completed: bool = False
    # Priority: "routine" = visible only inside patient page;
    #           "asap" = visible in dashboard "richieste urgenti" widget.
    priority: str = "routine"
    # Visibility: "shared" (default) = whole organization can see/edit;
    #             "private"           = only creator + shared_with_user_ids see it.
    visibility: str = "shared"
    shared_with_user_ids: List[str] = Field(default_factory=list)


class Reminder(ReminderBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    organization_id: str
    created_by: str
    created_by_name: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())


class ReminderUpdate(BaseModel):
    due_date: Optional[str] = None
    title: Optional[str] = None
    type: Optional[str] = None
    notes: Optional[str] = None
    completed: Optional[bool] = None
    priority: Optional[str] = None
    visibility: Optional[str] = None
    shared_with_user_ids: Optional[List[str]] = None


# ==================== PRO TOKEN (Patient-Reported Outcomes) ====================
# A PRO token is a short-lived signed link the doctor sends/prints/shows to a
# patient. The patient opens it on their phone and fills validated
# questionnaires (HAQ, BASDAI, BASFI, RAID, VAS pain/PGA, FIQR, PSAID, ESSPRI...).
# When submitted, results are stored on the token and the doctor reviews +
# approves them, optionally turning them into formal assessments.
PRO_INSTRUMENTS_ALLOWED = {
    "haq", "basdai", "basfi", "raid", "psaid", "fiqr", "esspri",
    "vas_pain", "vas_pga", "vas_fatigue",
}


class PROTokenCreate(BaseModel):
    patient_id: str
    instruments: List[str] = Field(default_factory=list)
    expires_in_hours: int = 168  # default 7 days
    note: Optional[str] = None  # optional note shown to the patient


class PROToken(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    token: str = Field(default_factory=lambda: secrets.token_urlsafe(24))
    patient_id: str
    organization_id: str
    instruments: List[str] = Field(default_factory=list)
    note: Optional[str] = None
    expires_at: str
    created_by: str
    created_by_name: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    completed_at: Optional[str] = None
    submitted_responses: Optional[Dict[str, Any]] = None
    # Indicates whether the responses were already converted to assessments.
    converted: bool = False


class PROSubmit(BaseModel):
    responses: Dict[str, Any]


# ==================== CONSULT TOKENS ====================
class ConsultTokenCreate(BaseModel):
    expires_in_hours: int = 168  # default 7 days


class ConsultToken(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    token: str = Field(default_factory=lambda: secrets.token_urlsafe(16))
    patient_id: str
    organization_id: str
    created_by: str
    created_by_name: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    expires_at: str
    views: int = 0


# ==================== SCLERODERMA PROFILE ====================
class ScleroProfileBase(BaseModel):
    patient_id: str
    cutaneous: Optional[Dict[str, Any]] = None       # subset, mrss, notes
    antibody: Optional[Dict[str, Any]] = None        # ana, aca, scl70, rnap3, ...
    vascular: Optional[Dict[str, Any]] = None        # raynaud, ulcers, capillaroscopy ...
    ild: Optional[Dict[str, Any]] = None             # presence, hrct, fvc, dlco
    pah: Optional[Dict[str, Any]] = None             # screen, echo, rhc
    gi: Optional[Dict[str, Any]] = None              # ger, sibo, dysmotility, ...
    msk: Optional[Dict[str, Any]] = None             # arthralgia, arthritis, contractures
    notes: Optional[str] = None


class ScleroProfile(ScleroProfileBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    organization_id: str
    created_by: str
    created_by_name: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    updated_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    updated_by_name: Optional[str] = None


# ==================== APP ====================
app = FastAPI(title="Clinimetria Reumatologica API")
api_router = APIRouter(prefix="/api")


# ==================== AUTH ENDPOINTS ====================
@api_router.post("/auth/register")
async def register(payload: RegisterRequest, response: Response):
    email = payload.email.strip().lower()
    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=400, detail="Email già registrata")

    # Determine organization
    org_id = None
    org_name = None
    role = "member"
    if payload.invite_code:
        org = await db.organizations.find_one({"invite_code": payload.invite_code.strip()}, {"_id": 0})
        if not org:
            raise HTTPException(status_code=400, detail="Codice invito non valido")
        org_id = org["id"]
        org_name = org["name"]
        role = "member"
    elif payload.organization_name:
        new_org = Organization(name=payload.organization_name.strip())
        await db.organizations.insert_one(new_org.model_dump())
        org_id = new_org.id
        org_name = new_org.name
        role = "admin"
    else:
        raise HTTPException(status_code=400, detail="Specificare nome dell'UO o codice invito")

    user_id = str(uuid.uuid4())
    user_doc = {
        "id": user_id,
        "email": email,
        "name": payload.name.strip(),
        "password_hash": hash_password(payload.password),
        "role": role,
        "organization_id": org_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.users.insert_one(user_doc)

    access = create_access_token(user_id, email, org_id)
    refresh = create_refresh_token(user_id)
    set_auth_cookies(response, access, refresh)

    return {
        "id": user_id, "email": email, "name": payload.name, "role": role,
        "organization_id": org_id, "organization_name": org_name,
    }


def _build_demo_seed_patients(days_ago) -> list:
    """Returns the 3 demo patients (RA, SpA, SLE) with realistic longitudinal data."""
    return [
        {
            "codice_paziente": "DEMO-AR-01",
            "cognome": "Bianchi", "nome": "Maria",
            "anno_nascita": 1968, "sesso": "F",
            "diagnosi": "Artrite reumatoide",
            "note": "RA sieropositiva erosiva, esordio 2019.",
            "assessments": [
                {"index_type": "das28_crp", "date": days_ago(540), "score": 5.6, "interpretation": "Alta attività", "tender": 8, "swollen": 6},
                {"index_type": "das28_crp", "date": days_ago(360), "score": 4.1, "interpretation": "Moderata attività", "tender": 5, "swollen": 3},
                {"index_type": "das28_crp", "date": days_ago(180), "score": 3.0, "interpretation": "Bassa attività", "tender": 2, "swollen": 1},
                {"index_type": "das28_crp", "date": days_ago(60), "score": 2.4, "interpretation": "Remissione", "tender": 1, "swollen": 0},
                {"index_type": "cdai", "date": days_ago(60), "score": 4, "interpretation": "Bassa attività"},
                {"index_type": "haq", "date": days_ago(60), "score": 0.5, "interpretation": "Disabilità lieve"},
            ],
            "therapies": [
                {"drug_name": "Methotrexate", "category": "csDMARD", "dose": "15 mg/sett", "frequency": "Settimanale", "route": "s.c.", "start_date": days_ago(540), "status": "active"},
                {"drug_name": "Prednisone", "category": "Glucocorticoidi", "dose": "5 mg", "frequency": "Giornaliera", "route": "oral", "start_date": days_ago(540), "end_date": days_ago(120), "status": "discontinued"},
                {"drug_name": "Adalimumab", "category": "bDMARD", "dose": "40 mg", "frequency": "Bisettimanale", "route": "s.c.", "start_date": days_ago(300), "status": "active"},
            ],
        },
        {
            "codice_paziente": "DEMO-SPA-02",
            "cognome": "Rossi", "nome": "Marco",
            "anno_nascita": 1985, "sesso": "M",
            "diagnosi": "Spondilite anchilosante",
            "note": "axSpA HLA-B27+, sacroileite RM positiva.",
            "assessments": [
                {"index_type": "basdai", "date": days_ago(365), "score": 6.8, "interpretation": "Alta attività"},
                {"index_type": "asdas_crp", "date": days_ago(365), "score": 3.4, "interpretation": "Alta attività"},
                {"index_type": "basdai", "date": days_ago(120), "score": 3.2, "interpretation": "Moderata"},
                {"index_type": "asdas_crp", "date": days_ago(120), "score": 1.9, "interpretation": "Bassa attività"},
                {"index_type": "basfi", "date": days_ago(120), "score": 2.5, "interpretation": "Funzione preservata"},
            ],
            "therapies": [
                {"drug_name": "Ibuprofene", "category": "FANS", "dose": "600 mg", "frequency": "TID al bisogno", "route": "oral", "start_date": days_ago(365), "status": "active"},
                {"drug_name": "Secukinumab", "category": "bDMARD", "dose": "150 mg", "frequency": "Mensile", "route": "s.c.", "start_date": days_ago(280), "status": "active"},
            ],
        },
        {
            "codice_paziente": "DEMO-SLE-03",
            "cognome": "Verdi", "nome": "Lucia",
            "anno_nascita": 1992, "sesso": "F",
            "diagnosi": "Lupus eritematoso sistemico (LES)",
            "note": "LES con coinvolgimento cutaneo, articolare, ANA+ a titolo elevato.",
            "assessments": [
                {"index_type": "sledai", "date": days_ago(420), "score": 12, "interpretation": "Alta attività"},
                {"index_type": "sledai", "date": days_ago(180), "score": 6, "interpretation": "Moderata"},
                {"index_type": "sledai", "date": days_ago(60), "score": 2, "interpretation": "Bassa attività"},
            ],
            "therapies": [
                {"drug_name": "Idrossiclorochina", "category": "csDMARD", "dose": "200 mg", "frequency": "BID", "route": "oral", "start_date": days_ago(420), "status": "active"},
                {"drug_name": "Prednisone", "category": "Glucocorticoidi", "dose": "10 mg", "frequency": "Giornaliera", "route": "oral", "start_date": days_ago(420), "end_date": days_ago(150), "status": "discontinued"},
                {"drug_name": "Belimumab", "category": "bDMARD", "dose": "200 mg", "frequency": "Settimanale", "route": "s.c.", "start_date": days_ago(180), "status": "active"},
            ],
        },
    ]


async def _insert_demo_patient(p_seed: dict, org_id: str, user_id: str) -> None:
    """Inserts one demo patient with its assessments and therapies."""
    p_id = str(uuid.uuid4())
    now_iso = datetime.now(timezone.utc).isoformat()
    audit = {"created_at": now_iso, "created_by": user_id, "created_by_name": "Utente Demo"}

    await db.patients.insert_one({
        "id": p_id, "organization_id": org_id,
        "codice_paziente": p_seed["codice_paziente"],
        "cognome": p_seed["cognome"], "nome": p_seed["nome"],
        "anno_nascita": p_seed["anno_nascita"], "sesso": p_seed["sesso"],
        "diagnosi": p_seed["diagnosi"], "note": p_seed["note"],
        **audit,
    })
    for a in p_seed["assessments"]:
        await db.assessments.insert_one({
            "id": str(uuid.uuid4()), "patient_id": p_id, "organization_id": org_id,
            "index_type": a["index_type"], "date": a["date"], "score": a["score"],
            "interpretation": a.get("interpretation"),
            "tender_joints": [f"j{i}" for i in range(a.get("tender", 0))],
            "swollen_joints": [f"j{i}" for i in range(a.get("swollen", 0))],
            "inputs": {}, **audit,
        })
    for t in p_seed["therapies"]:
        await db.therapies.insert_one({
            "id": str(uuid.uuid4()), "patient_id": p_id, "organization_id": org_id,
            "drug_name": t["drug_name"], "category": t["category"],
            "dose": t.get("dose"), "frequency": t.get("frequency"), "route": t.get("route"),
            "start_date": t["start_date"], "end_date": t.get("end_date"),
            "status": t["status"], "indication": p_seed["diagnosi"],
            **audit,
        })


@api_router.post("/auth/demo")
async def login_demo(response: Response):
    """
    One-click demo: creates a fresh isolated organization with 3 pre-populated
    patients (RA, SpA, SLE) with assessments and therapies, and signs the user
    in immediately. Each call creates a NEW demo org (so multiple users can
    explore in parallel).
    """
    import secrets
    suffix = secrets.token_hex(4)  # 8 lowercase hex chars; cryptographically strong
    org = Organization(name=f"Demo UO {suffix}")
    await db.organizations.insert_one(org.model_dump())

    user_id = str(uuid.uuid4())
    email = f"demo-{suffix}@clinimetria.demo"
    await db.users.insert_one({
        "id": user_id, "email": email, "name": "Utente Demo",
        "password_hash": hash_password(f"demo-{suffix}"),
        "role": "admin", "organization_id": org.id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "is_demo": True,
    })

    today = datetime.now(timezone.utc).date()
    def days_ago(n: int) -> str:
        return (today - timedelta(days=n)).isoformat()

    for p_seed in _build_demo_seed_patients(days_ago):
        await _insert_demo_patient(p_seed, org.id, user_id)

    access = create_access_token(user_id, email, org.id)
    refresh = create_refresh_token(user_id)
    set_auth_cookies(response, access, refresh)
    return {
        "id": user_id, "email": email, "name": "Utente Demo", "role": "admin",
        "organization_id": org.id, "organization_name": org.name,
        "is_demo": True,
    }


@api_router.post("/auth/login")
async def login(payload: LoginRequest, response: Response):
    email = payload.email.strip().lower()
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(payload.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Credenziali non valide")

    org = await db.organizations.find_one({"id": user["organization_id"]}, {"_id": 0})
    org_name = org["name"] if org else None

    access = create_access_token(user["id"], user["email"], user["organization_id"])
    refresh = create_refresh_token(user["id"])
    set_auth_cookies(response, access, refresh)

    return {
        "id": user["id"], "email": user["email"], "name": user["name"], "role": user["role"],
        "organization_id": user["organization_id"], "organization_name": org_name,
    }


@api_router.post("/auth/logout")
async def logout(response: Response, _: dict = Depends(get_current_user)):
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    return {"success": True}


@api_router.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    org = await db.organizations.find_one({"id": user["organization_id"]}, {"_id": 0})
    return {
        "id": user["id"], "email": user["email"], "name": user["name"], "role": user.get("role", "member"),
        "organization_id": user["organization_id"], "organization_name": org["name"] if org else None,
        "invite_code": org["invite_code"] if org else None,
        "pseudonymized_mode": bool(org.get("pseudonymized_mode", False)) if org else False,
        "ai_strict_privacy": bool(org.get("ai_strict_privacy", False)) if org else False,
        "is_demo": bool(user.get("is_demo", False)),
    }


@api_router.put("/organization/settings")
async def update_org_settings(payload: OrganizationSettings, user: dict = Depends(get_current_user)):
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Solo gli amministratori possono modificare le impostazioni")
    update = {k: v for k, v in payload.model_dump().items() if v is not None}
    if not update:
        raise HTTPException(status_code=400, detail="Nessuna modifica")
    await db.organizations.update_one({"id": user["organization_id"]}, {"$set": update})
    return {"success": True, **update}


@api_router.post("/auth/refresh")
async def refresh_token(request: Request, response: Response):
    token = request.cookies.get("refresh_token")
    if not token:
        raise HTTPException(status_code=401, detail="Token di refresh assente")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "refresh":
            raise HTTPException(status_code=401, detail="Tipo di token non valido")
        user = await db.users.find_one({"id": payload["sub"]})
        if not user:
            raise HTTPException(status_code=401, detail="Utente non trovato")
        access = create_access_token(user["id"], user["email"], user["organization_id"])
        response.set_cookie(key="access_token", value=access, httponly=True, secure=True, samesite="none", max_age=28800, path="/")
        return {"success": True}
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Token non valido")


# ==================== PATIENTS (auth + org-scoped) ====================
@api_router.get("/")
async def root():
    return {"message": "Clinimetria Reumatologica API"}


@api_router.post("/patients", response_model=Patient)
async def create_patient(payload: PatientBase, user: dict = Depends(get_current_user)):
    # Require at least codice_paziente OR (nome + cognome) to identify the patient
    has_code = bool((payload.codice_paziente or "").strip())
    has_name = bool((payload.nome or "").strip() and (payload.cognome or "").strip())
    if not has_code and not has_name:
        raise HTTPException(
            status_code=400,
            detail="Fornire almeno il codice paziente o nome+cognome",
        )
    # Enforce uniqueness of codice_paziente within the org
    if has_code:
        existing = await db.patients.find_one({
            "organization_id": user["organization_id"],
            "codice_paziente": payload.codice_paziente.strip(),
        })
        if existing:
            raise HTTPException(status_code=400, detail="Codice paziente già esistente nell'UO")
    p = Patient(**payload.model_dump(), organization_id=user["organization_id"], created_by=user["id"], created_by_name=user.get("name"))
    await db.patients.insert_one(p.model_dump())
    return p


@api_router.post("/patients/{patient_id}/anonymize", response_model=Patient)
async def anonymize_patient(patient_id: str, user: dict = Depends(get_current_user)):
    """Rimuove i dati identificativi (nome, cognome, CF, data_nascita) lasciando solo
    codice paziente + anno di nascita + sesso + diagnosi."""
    await _verify_patient_in_org(patient_id, user["organization_id"])
    patient = await db.patients.find_one(
        {"id": patient_id, "organization_id": user["organization_id"]}, {"_id": 0}
    )
    if not patient:
        raise HTTPException(status_code=404, detail="Paziente non trovato")

    # Compute anno_nascita from data_nascita if present and not already set
    update = {"nome": None, "cognome": None, "codice_fiscale": None, "data_nascita": None}
    if patient.get("data_nascita") and not patient.get("anno_nascita"):
        try:
            update["anno_nascita"] = int(patient["data_nascita"][:4])
        except (ValueError, TypeError):
            pass
    # Ensure codice_paziente exists; if not, generate one
    if not patient.get("codice_paziente"):
        update["codice_paziente"] = f"PZ-{patient_id[:8].upper()}"
    await db.patients.update_one({"id": patient_id}, {"$set": update})
    return await db.patients.find_one({"id": patient_id}, {"_id": 0})


@api_router.get("/patients", response_model=List[Patient])
async def list_patients(user: dict = Depends(get_current_user)):
    docs = await db.patients.find({"organization_id": user["organization_id"]}, {"_id": 0}).to_list(2000)
    # Sort in Python to tolerate missing cognome (pseudonymized patients)
    docs.sort(key=lambda p: (p.get("cognome") or p.get("codice_paziente") or "").lower())
    return docs


@api_router.get("/patients/{patient_id}", response_model=Patient)
async def get_patient(patient_id: str, user: dict = Depends(get_current_user)):
    doc = await db.patients.find_one({"id": patient_id, "organization_id": user["organization_id"]}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Paziente non trovato")
    return doc


@api_router.put("/patients/{patient_id}", response_model=Patient)
async def update_patient(patient_id: str, payload: PatientUpdate, user: dict = Depends(get_current_user)):
    update_data = {k: v for k, v in payload.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Nessun campo da aggiornare")
    result = await db.patients.update_one(
        {"id": patient_id, "organization_id": user["organization_id"]},
        {"$set": update_data},
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Paziente non trovato")
    return await db.patients.find_one({"id": patient_id}, {"_id": 0})


@api_router.delete("/patients/{patient_id}")
async def delete_patient(patient_id: str, user: dict = Depends(get_current_user)):
    result = await db.patients.delete_one({"id": patient_id, "organization_id": user["organization_id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Paziente non trovato")
    await db.assessments.delete_many({"patient_id": patient_id})
    await db.criteria_evaluations.delete_many({"patient_id": patient_id})
    await db.therapies.delete_many({"patient_id": patient_id})
    await db.lab_exams.delete_many({"patient_id": patient_id})
    await db.reminders.delete_many({"patient_id": patient_id})
    await db.sclero_profiles.delete_many({"patient_id": patient_id})
    await db.disease_profiles.delete_many({"patient_id": patient_id})
    return {"success": True}


# ==================== RECALL FLAGS (pazienti da ricontrollare) ====================
# Each patient can have at most one recall flag, with four states:
#   - flag = "private"  → visibile solo all'autore (stellina BLU)
#   - flag = "selected" → visibile all'autore + colleghi scelti (stellina GIALLA)
#   - flag = "shared"   → visibile a tutti i membri dell'organizzazione (stellina ROSSA)
#   - flag = None       → nessun ricontrollo
# Stored as `patients.recall = {flag, note, set_at, set_by, set_by_name, shared_with}`.
# `shared_with` is a list of user IDs; only used when flag == "selected".

class RecallFlagPayload(BaseModel):
    flag: Optional[str] = None  # "private" | "selected" | "shared" | None
    note: Optional[str] = None
    shared_with: Optional[List[str]] = None  # user IDs; only for flag == "selected"


@api_router.put("/patients/{patient_id}/recall")
async def set_patient_recall(
    patient_id: str,
    payload: RecallFlagPayload,
    user: dict = Depends(get_current_user),
):
    if payload.flag not in (None, "", "private", "selected", "shared"):
        raise HTTPException(status_code=400, detail="flag deve essere 'private', 'selected', 'shared' o nullo")
    p = await db.patients.find_one({"id": patient_id, "organization_id": user["organization_id"]}, {"_id": 0, "id": 1})
    if not p:
        raise HTTPException(status_code=404, detail="Paziente non trovato")

    if not payload.flag:
        await db.patients.update_one({"id": patient_id}, {"$unset": {"recall": ""}})
        return {"success": True, "recall": None}

    recall = {
        "flag": payload.flag,
        "note": (payload.note or "").strip(),
        "set_at": datetime.now(timezone.utc).isoformat(),
        "set_by": user["id"],
        "set_by_name": user.get("name") or user.get("email"),
    }
    if payload.flag == "selected":
        # Store the explicit list of colleague IDs (excluding the setter themselves)
        recall["shared_with"] = [uid for uid in (payload.shared_with or []) if uid != user["id"]]
    else:
        recall["shared_with"] = []

    await db.patients.update_one({"id": patient_id}, {"$set": {"recall": recall}})
    return {"success": True, "recall": recall}


@api_router.get("/patients-recall")
async def list_patients_to_recall(user: dict = Depends(get_current_user)):
    """Pazienti con recall flag visibili a questo utente:
    - tutti i 'shared' dell'org
    - solo i 'private' che ho creato io
    - i 'selected' che ho creato io, oppure in cui sono nella lista shared_with
    """
    org_id = user["organization_id"]
    uid = user["id"]
    cursor = db.patients.find(
        {
            "organization_id": org_id,
            "$or": [
                {"recall.flag": "shared"},
                {"recall.flag": "private", "recall.set_by": uid},
                {"recall.flag": "selected", "recall.set_by": uid},
                {"recall.flag": "selected", "recall.shared_with": uid},
            ],
        },
        {"_id": 0},
    )
    out = []
    async for p in cursor:
        out.append(p)
    out.sort(key=lambda p: (p.get("recall", {}).get("set_at") or ""), reverse=True)
    return out


@api_router.get("/patients-recent-mine")
async def list_recent_patients_mine(days: int = 7, user: dict = Depends(get_current_user)):
    """Pazienti su cui IO ho lavorato (creato/aggiornato un assessment) negli ultimi `days` giorni."""
    org_id = user["organization_id"]
    if days < 1:
        days = 7
    if days > 365:
        days = 365
    cutoff = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()

    pipeline = [
        {"$match": {
            "organization_id": org_id,
            "created_by": user["id"],
            "created_at": {"$gte": cutoff},
        }},
        {"$sort": {"created_at": -1}},
        {"$group": {
            "_id": "$patient_id",
            "last_assessment_at": {"$first": "$created_at"},
            "last_index_type": {"$first": "$index_type"},
            "last_score": {"$first": "$score"},
            "count": {"$sum": 1},
        }},
        {"$sort": {"last_assessment_at": -1}},
        {"$limit": 20},
    ]
    items = []
    async for doc in db.assessments.aggregate(pipeline):
        items.append(doc)
    if not items:
        return []
    pids = [i["_id"] for i in items]
    patients = {
        p["id"]: p
        for p in await db.patients.find({"id": {"$in": pids}, "organization_id": org_id}, {"_id": 0}).to_list(1000)
    }
    out = []
    for it in items:
        p = patients.get(it["_id"])
        if not p:
            continue
        out.append({
            **p,
            "last_assessment_at": it["last_assessment_at"],
            "last_index_type": it["last_index_type"],
            "last_score": it["last_score"],
            "assessments_in_window": it["count"],
        })
    return out


# ==================== ASSESSMENTS ====================
async def _verify_patient_in_org(patient_id: str, organization_id: str):
    p = await db.patients.find_one({"id": patient_id, "organization_id": organization_id}, {"_id": 0, "id": 1})
    if not p:
        raise HTTPException(status_code=404, detail="Paziente non trovato")


@api_router.post("/assessments", response_model=Assessment)
async def create_assessment(payload: AssessmentBase, user: dict = Depends(get_current_user)):
    await _verify_patient_in_org(payload.patient_id, user["organization_id"])
    a = Assessment(**payload.model_dump(), organization_id=user["organization_id"], created_by=user["id"], created_by_name=user.get("name"))
    await db.assessments.insert_one(a.model_dump())
    return a


@api_router.put("/assessments/{assessment_id}", response_model=Assessment)
async def update_assessment(assessment_id: str, payload: AssessmentBase, user: dict = Depends(get_current_user)):
    update_data = payload.model_dump()
    result = await db.assessments.update_one(
        {"id": assessment_id, "organization_id": user["organization_id"]},
        {"$set": update_data},
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Valutazione non trovata")
    return await db.assessments.find_one({"id": assessment_id}, {"_id": 0})


@api_router.get("/patients/{patient_id}/assessments", response_model=List[Assessment])
async def list_patient_assessments(patient_id: str, user: dict = Depends(get_current_user)):
    await _verify_patient_in_org(patient_id, user["organization_id"])
    docs = await db.assessments.find({"patient_id": patient_id}, {"_id": 0}).sort("date", -1).to_list(2000)
    return docs


@api_router.get("/assessments/{assessment_id}", response_model=Assessment)
async def get_assessment(assessment_id: str, user: dict = Depends(get_current_user)):
    doc = await db.assessments.find_one({"id": assessment_id, "organization_id": user["organization_id"]}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Valutazione non trovata")
    return doc


@api_router.delete("/assessments/{assessment_id}")
async def delete_assessment(assessment_id: str, user: dict = Depends(get_current_user)):
    result = await db.assessments.delete_one({"id": assessment_id, "organization_id": user["organization_id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Valutazione non trovata")
    return {"success": True}


# ==================== CRITERIA EVALUATIONS ====================
@api_router.post("/criteria-evaluations", response_model=CriteriaEvaluation)
async def create_criteria_evaluation(payload: CriteriaEvaluationBase, user: dict = Depends(get_current_user)):
    await _verify_patient_in_org(payload.patient_id, user["organization_id"])
    ev = CriteriaEvaluation(**payload.model_dump(), organization_id=user["organization_id"], created_by=user["id"], created_by_name=user.get("name"))
    await db.criteria_evaluations.insert_one(ev.model_dump())
    return ev


@api_router.get("/patients/{patient_id}/criteria-evaluations", response_model=List[CriteriaEvaluation])
async def list_patient_criteria(patient_id: str, user: dict = Depends(get_current_user)):
    await _verify_patient_in_org(patient_id, user["organization_id"])
    docs = await db.criteria_evaluations.find({"patient_id": patient_id}, {"_id": 0}).sort("date", -1).to_list(2000)
    return docs


@api_router.delete("/criteria-evaluations/{evaluation_id}")
async def delete_criteria_evaluation(evaluation_id: str, user: dict = Depends(get_current_user)):
    result = await db.criteria_evaluations.delete_one({"id": evaluation_id, "organization_id": user["organization_id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Valutazione non trovata")
    return {"success": True}


# ==================== THERAPIES ====================
# Categorie che richiedono esclusione (un solo farmaco di queste categorie attivo per volta)
EXCLUSIVE_CATEGORIES = {"bDMARD", "tsDMARD"}


def _extract_dose_mg(dose_str: Optional[str]) -> Optional[float]:
    """Extract the numeric mg value from a free-text dose string, or None."""
    if not dose_str:
        return None
    m = re.search(r"(\d+(?:[.,]\d+)?)\s*mg", dose_str, re.IGNORECASE)
    return float(m.group(1).replace(",", ".")) if m else None


def _therapy_event_projection(event: dict) -> dict:
    """Project the current therapy fields implied by a non-voided therapy event."""
    if event.get("voided"):
        return {}

    event_type = event.get("type")
    set_fields: dict = {}
    if event_type == "discontinued":
        set_fields["status"] = event.get("status_after") or "discontinued"
        set_fields["end_date"] = event.get("date")
        if event.get("reason"):
            set_fields["discontinuation_reason"] = event.get("reason")
    elif event_type in ("dose_increased", "dose_reduced", "regimen_changed"):
        if event.get("dose_after"):
            set_fields["dose"] = event.get("dose_after")
        if event.get("frequency_after"):
            set_fields["frequency"] = event.get("frequency_after")
        if event.get("route_after"):
            set_fields["route"] = event.get("route_after")
    elif event_type == "resumed_within":
        set_fields["status"] = "active"
        set_fields["end_date"] = None
        if event.get("dose"):
            set_fields["dose"] = event.get("dose")
    elif event_type == "status_changed" and event.get("status_after"):
        set_fields["status"] = event.get("status_after")

    return set_fields


async def _append_therapy_event_and_project(
    therapy_filter: dict,
    event: dict,
    extra_set: Optional[dict] = None,
):
    """
    Append a therapy event and project event-derived current fields atomically.
    All writes to therapy longitudinal fields must go through this helper.
    """
    set_fields = dict(extra_set or {})
    set_fields.update(_therapy_event_projection(event))

    update_op: dict = {"$push": {"events": event}}
    if set_fields:
        update_op["$set"] = set_fields

    return await db.therapies.update_one(therapy_filter, update_op)


async def _auto_discontinue_competing(patient_id: str, organization_id: str, new_category: str, new_start_date: Optional[str], exclude_id: Optional[str] = None) -> int:
    """If new therapy is in EXCLUSIVE_CATEGORIES, append discontinued events to other active biologic/tsDMARD therapies."""
    if new_category not in EXCLUSIVE_CATEGORIES:
        return 0
    end_date = new_start_date or datetime.now(timezone.utc).date().isoformat()
    query = {
        "patient_id": patient_id,
        "organization_id": organization_id,
        "category": {"$in": list(EXCLUSIVE_CATEGORIES)},
        "status": "active",
    }
    if exclude_id:
        query["id"] = {"$ne": exclude_id}
    modified = 0
    async for therapy in db.therapies.find(query, {"_id": 0, "id": 1, "status": 1, "dose": 1}):
        event = TherapyEvent(
            type="discontinued",
            date=end_date,
            dose=therapy.get("dose"),
            status_before=therapy.get("status"),
            status_after="discontinued",
            reason="auto_discontinued_competing_therapy",
        ).model_dump()
        result = await _append_therapy_event_and_project(
            {"id": therapy["id"], "organization_id": organization_id},
            event,
            {"auto_discontinued": True},
        )
        modified += result.modified_count
    return modified


@api_router.post("/therapies", response_model=Therapy)
async def create_therapy(payload: TherapyBase, user: dict = Depends(get_current_user)):
    await _verify_patient_in_org(payload.patient_id, user["organization_id"])
    # Apply exclusivity rule for biologics/tsDMARDs BEFORE inserting the new one
    if (payload.status or "active") == "active":
        await _auto_discontinue_competing(
            payload.patient_id, user["organization_id"], payload.category, payload.start_date
        )

    data = payload.model_dump()
    if not data.get("events"):
        event_type = "started" if (payload.status or "active") == "active" else "discontinued"
        data["events"] = [TherapyEvent(
            type=event_type,
            date=payload.start_date or payload.end_date or datetime.now(timezone.utc).date().isoformat(),
            dose=payload.dose,
            frequency_after=payload.frequency,
            route_after=payload.route,
            status_after=payload.status,
            reason=payload.discontinuation_reason if event_type == "discontinued" else None,
            created_by=user["id"],
            created_by_name=user.get("name"),
            notes=payload.notes,
        ).model_dump()]

    t = Therapy(**data, organization_id=user["organization_id"], created_by=user["id"], created_by_name=user.get("name"))
    await db.therapies.insert_one(t.model_dump())
    return t


@api_router.get("/patients/{patient_id}/therapies", response_model=List[Therapy])
async def list_patient_therapies(patient_id: str, user: dict = Depends(get_current_user)):
    await _verify_patient_in_org(patient_id, user["organization_id"])
    docs = await db.therapies.find({"patient_id": patient_id}, {"_id": 0}).sort("start_date", -1).to_list(2000)
    return docs


@api_router.put("/therapies/{therapy_id}", response_model=Therapy)
async def update_therapy(therapy_id: str, payload: TherapyUpdate, user: dict = Depends(get_current_user)):
    update_data = {k: v for k, v in payload.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Nessun campo da aggiornare")
    # Fetch current therapy to know patient/category
    current = await db.therapies.find_one({"id": therapy_id, "organization_id": user["organization_id"]}, {"_id": 0})
    if not current:
        raise HTTPException(status_code=404, detail="Terapia non trovata")

    longitudinal_fields = {
        "status",
        "dose",
        "frequency",
        "route",
        "end_date",
        "discontinuation_reason",
    }
    direct_update = {k: v for k, v in update_data.items() if k not in longitudinal_fields}
    longitudinal_update = {k: v for k, v in update_data.items() if k in longitudinal_fields}

    # Apply exclusivity if the (resulting) therapy is an active biologic/tsDMARD
    new_status = update_data.get("status", current.get("status"))
    new_category = update_data.get("category", current.get("category"))
    new_start = update_data.get("start_date", current.get("start_date"))
    if new_status == "active":
        await _auto_discontinue_competing(
            current["patient_id"], user["organization_id"], new_category, new_start, exclude_id=therapy_id
        )

    if longitudinal_update:
        today = datetime.now(timezone.utc).date().isoformat()
        event_date = longitudinal_update.get("end_date") or today
        reason = longitudinal_update.get("discontinuation_reason") or update_data.get("notes")
        status_after = longitudinal_update.get("status")
        event_type = None

        if status_after == "active" and current.get("status") != "active":
            event_type = "resumed_within"
        elif status_after in {"discontinued", "completed"} or "end_date" in longitudinal_update or "discontinuation_reason" in longitudinal_update:
            event_type = "discontinued"
        elif "dose" in longitudinal_update:
            current_mg = _extract_dose_mg(current.get("dose"))
            new_mg = _extract_dose_mg(longitudinal_update.get("dose"))
            if current_mg is not None and new_mg is not None and new_mg != current_mg:
                event_type = "dose_increased" if new_mg > current_mg else "dose_reduced"
            else:
                event_type = "regimen_changed"
        elif "frequency" in longitudinal_update or "route" in longitudinal_update:
            event_type = "regimen_changed"
        elif status_after:
            event_type = "status_changed"

        event = TherapyEvent(
            type=event_type or "status_changed",
            date=event_date,
            dose=longitudinal_update.get("dose", current.get("dose")),
            dose_before=current.get("dose") if "dose" in longitudinal_update else None,
            dose_after=longitudinal_update.get("dose") if "dose" in longitudinal_update else None,
            frequency_before=current.get("frequency") if "frequency" in longitudinal_update else None,
            frequency_after=longitudinal_update.get("frequency") if "frequency" in longitudinal_update else None,
            route_before=current.get("route") if "route" in longitudinal_update else None,
            route_after=longitudinal_update.get("route") if "route" in longitudinal_update else None,
            status_before=current.get("status") if "status" in longitudinal_update else None,
            status_after=status_after,
            reason=reason,
            created_by=user["id"],
            created_by_name=user.get("name"),
            notes=update_data.get("notes"),
        ).model_dump()
        result = await _append_therapy_event_and_project(
            {"id": therapy_id, "organization_id": user["organization_id"]},
            event,
            direct_update,
        )
    elif direct_update:
        result = await db.therapies.update_one(
            {"id": therapy_id, "organization_id": user["organization_id"]},
            {"$set": direct_update},
        )
    else:
        result = None

    if result and result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Terapia non trovata")
    return await db.therapies.find_one({"id": therapy_id, "organization_id": user["organization_id"]}, {"_id": 0})


@api_router.delete("/therapies/{therapy_id}")
async def delete_therapy(therapy_id: str, user: dict = Depends(get_current_user)):
    result = await db.therapies.delete_one({"id": therapy_id, "organization_id": user["organization_id"]})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Terapia non trovata")
    return {"success": True}


# ==================== LAB EXAMS ====================
@api_router.post("/lab-exams", response_model=LabExam)
async def create_lab_exam(payload: LabExamBase, user: dict = Depends(get_current_user)):
    await _verify_patient_in_org(payload.patient_id, user["organization_id"])
    e = LabExam(**payload.model_dump(), organization_id=user["organization_id"], created_by=user["id"], created_by_name=user.get("name"))
    await db.lab_exams.insert_one(e.model_dump())
    return e


@api_router.get("/patients/{patient_id}/lab-exams", response_model=List[LabExam])
async def list_patient_lab_exams(patient_id: str, user: dict = Depends(get_current_user)):
    await _verify_patient_in_org(patient_id, user["organization_id"])
    docs = await db.lab_exams.find({"patient_id": patient_id}, {"_id": 0}).sort("date", -1).to_list(2000)
    return docs


@api_router.put("/lab-exams/{exam_id}", response_model=LabExam)
async def update_lab_exam(exam_id: str, payload: LabExamBase, user: dict = Depends(get_current_user)):
    update_data = payload.model_dump()
    result = await db.lab_exams.update_one(
        {"id": exam_id, "organization_id": user["organization_id"]},
        {"$set": update_data},
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Esame non trovato")
    return await db.lab_exams.find_one({"id": exam_id}, {"_id": 0})


@api_router.delete("/lab-exams/{exam_id}")
async def delete_lab_exam(exam_id: str, user: dict = Depends(get_current_user)):
    result = await db.lab_exams.delete_one({"id": exam_id, "organization_id": user["organization_id"]})


# ==================== SPECIALIST VISITS ====================
@api_router.post("/specialist-visits", response_model=SpecialistVisit)
async def create_specialist_visit(payload: SpecialistVisitBase, user: dict = Depends(get_current_user)):
    await _verify_patient_in_org(payload.patient_id, user["organization_id"])
    sv = SpecialistVisit(
        **payload.model_dump(),
        organization_id=user["organization_id"],
        created_by=user["id"],
        created_by_name=user.get("name") or user.get("email"),
    )
    await db.specialist_visits.insert_one(sv.model_dump())
    return sv


@api_router.get("/patients/{patient_id}/specialist-visits", response_model=List[SpecialistVisit])
async def list_specialist_visits(patient_id: str, user: dict = Depends(get_current_user)):
    await _verify_patient_in_org(patient_id, user["organization_id"])
    docs = await db.specialist_visits.find(
        {"patient_id": patient_id, "organization_id": user["organization_id"]},
        {"_id": 0},
    ).sort("visit_date", -1).to_list(500)
    return docs


@api_router.put("/specialist-visits/{visit_id}", response_model=SpecialistVisit)
async def update_specialist_visit(visit_id: str, payload: SpecialistVisitBase, user: dict = Depends(get_current_user)):
    await db.specialist_visits.update_one(
        {"id": visit_id, "organization_id": user["organization_id"]},
        {"$set": payload.model_dump()},
    )
    return await db.specialist_visits.find_one({"id": visit_id}, {"_id": 0})


@api_router.delete("/specialist-visits/{visit_id}")
async def delete_specialist_visit(visit_id: str, user: dict = Depends(get_current_user)):
    result = await db.specialist_visits.delete_one(
        {"id": visit_id, "organization_id": user["organization_id"]}
    )
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Visita non trovata")
    return {"success": True}


# ==================== REMINDERS ====================
def _reminder_visibility_query(user: dict) -> dict:
    """Filter clause: show shared org reminders OR private ones owned/shared with current user."""
    return {
        "$or": [
            {"visibility": {"$ne": "private"}},
            {"created_by": user["id"]},
            {"shared_with_user_ids": user["id"]},
        ]
    }


@api_router.post("/reminders", response_model=Reminder)
async def create_reminder(payload: ReminderBase, user: dict = Depends(get_current_user)):
    if payload.patient_id:
        await _verify_patient_in_org(payload.patient_id, user["organization_id"])
    r = Reminder(**payload.model_dump(), organization_id=user["organization_id"], created_by=user["id"], created_by_name=user.get("name"))
    await db.reminders.insert_one(r.model_dump())
    return r


@api_router.get("/patients/{patient_id}/reminders", response_model=List[Reminder])
async def list_patient_reminders(patient_id: str, user: dict = Depends(get_current_user)):
    await _verify_patient_in_org(patient_id, user["organization_id"])
    q = {
        "patient_id": patient_id,
        "organization_id": user["organization_id"],
        **_reminder_visibility_query(user),
    }
    docs = await db.reminders.find(q, {"_id": 0}).sort("due_date", 1).to_list(2000)
    return docs


@api_router.get("/reminders/upcoming", response_model=List[Reminder])
async def upcoming_reminders(user: dict = Depends(get_current_user)):
    """Dashboard widget: ALL non-completed tasks visible to user.
    Standalone tasks (no patient) always appear; patient-linked tasks also included.
    Sorted by due_date ascending (overdue first), limit 100."""
    q = {
        "organization_id": user["organization_id"],
        "completed": False,
        **_reminder_visibility_query(user),
    }
    docs = await db.reminders.find(q, {"_id": 0}).sort("due_date", 1).limit(100).to_list(100)
    return docs


@api_router.put("/reminders/{reminder_id}", response_model=Reminder)
async def update_reminder(reminder_id: str, payload: ReminderUpdate, user: dict = Depends(get_current_user)):
    update_data = {k: v for k, v in payload.model_dump().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="Nessun campo da aggiornare")
    # Visibility check: ensure user can update this reminder
    existing = await db.reminders.find_one(
        {"id": reminder_id, "organization_id": user["organization_id"]},
        {"_id": 0},
    )
    if not existing:
        raise HTTPException(status_code=404, detail="Reminder non trovato")
    if existing.get("visibility") == "private" and existing.get("created_by") != user["id"] and user["id"] not in (existing.get("shared_with_user_ids") or []):
        raise HTTPException(status_code=403, detail="Non hai accesso a questa richiesta privata")
    await db.reminders.update_one(
        {"id": reminder_id, "organization_id": user["organization_id"]},
        {"$set": update_data},
    )
    return await db.reminders.find_one({"id": reminder_id}, {"_id": 0})


@api_router.delete("/reminders/{reminder_id}")
async def delete_reminder(reminder_id: str, user: dict = Depends(get_current_user)):
    existing = await db.reminders.find_one(
        {"id": reminder_id, "organization_id": user["organization_id"]},
        {"_id": 0},
    )
    if not existing:
        raise HTTPException(status_code=404, detail="Reminder non trovato")
    if existing.get("visibility") == "private" and existing.get("created_by") != user["id"]:
        raise HTTPException(status_code=403, detail="Solo il creatore può eliminare una richiesta privata")
    await db.reminders.delete_one({"id": reminder_id, "organization_id": user["organization_id"]})
    return {"success": True}


# ==================== VISIT TEMPLATES ====================

class VisitTemplateBase(BaseModel):
    category: str   # "rheumatic_history" | "physical_exam"
    name: str
    content: str

class VisitTemplate(VisitTemplateBase):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    organization_id: str
    created_at: str = Field(default_factory=lambda: datetime.utcnow().isoformat())
    updated_at: str = Field(default_factory=lambda: datetime.utcnow().isoformat())

class VisitTemplateUpdate(BaseModel):
    name: Optional[str] = None
    content: Optional[str] = None

@api_router.get("/visit-templates", response_model=List[VisitTemplate])
async def list_visit_templates(category: Optional[str] = None, user: dict = Depends(get_current_user)):
    query: Dict[str, Any] = {"user_id": user["id"]}
    if category:
        query["category"] = category
    docs = await db.visit_templates.find(query, {"_id": 0}).sort("name", 1).to_list(500)
    return docs

@api_router.post("/visit-templates", response_model=VisitTemplate)
async def create_visit_template(payload: VisitTemplateBase, user: dict = Depends(get_current_user)):
    tpl = VisitTemplate(**payload.model_dump(), user_id=user["id"], organization_id=user["organization_id"])
    await db.visit_templates.insert_one(tpl.model_dump())
    return tpl

@api_router.put("/visit-templates/{tpl_id}", response_model=VisitTemplate)
async def update_visit_template(tpl_id: str, payload: VisitTemplateUpdate, user: dict = Depends(get_current_user)):
    existing = await db.visit_templates.find_one({"id": tpl_id, "user_id": user["id"]}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Template non trovato")
    update_data = {k: v for k, v in payload.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.utcnow().isoformat()
    await db.visit_templates.update_one({"id": tpl_id}, {"$set": update_data})
    return {**existing, **update_data}

@api_router.delete("/visit-templates/{tpl_id}")
async def delete_visit_template(tpl_id: str, user: dict = Depends(get_current_user)):
    existing = await db.visit_templates.find_one({"id": tpl_id, "user_id": user["id"]}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Template non trovato")
    await db.visit_templates.delete_one({"id": tpl_id})
    return {"success": True}


# ==================== WORKUP VISITS ====================

class WorkupVisitBase(BaseModel):
    patient_id: str
    visit_date: str                                   # ISO "YYYY-MM-DD"
    # ── Structured visit sections (canonical 7-section form) ──────────────────
    rheumatologic_history_summary: Optional[str] = None  # 0 · Raccordo anamnestico reumatologico
    interval_history: Optional[str] = None            # 1 · Anamnesi intervallare
    physical_exam: Optional[str] = None               # 2 · Esame obiettivo (free text)
    physical_exam_joint_exam: Optional[dict] = None  # 2 · Esame articolare (homunculus map)
    physical_exam_systems: Optional[dict] = None      # 2 · Apparati per sistema (organ fields)
    physical_exam_mrss: Optional[dict] = None        # 2 · MRSS (sclerosi sistemica)
    physical_exam_pasi: Optional[dict] = None        # 2 · PASI (psoriasi)
    physical_exam_lei: Optional[dict] = None         # 2 · LEI (entesiti)
    labs_imaging: Optional[str] = None                # 3 · Esami / imaging in visione
    clinimetria_notes: Optional[str] = None           # 4 · Clinimetria (opzionale)
    diagnostic_hypotheses: Optional[str] = None       # 5 · Assessment: ipotesi diagnostiche
    conclusions: Optional[str] = None                 # 5 · Assessment: conclusioni
    clinical_decision: str = "open"                   # "open" | "converting"
    confirmed_diagnosis: Optional[str] = None         # 5 · Assessment: diagnosi confermata (when converting)
    requested_tests: Optional[List[str]] = None       # 6 · Piano: esami richiesti
    requested_tests_notes: Optional[str] = None       # 6 · Piano: note esami richiesti
    followup_date: Optional[str] = None               # 6 · Piano: data prossima rivalutazione
    therapy_modification: Optional[str] = None        # 6 · Piano: modifica / indicazioni terapeutiche
    referred_to_gp: Optional[bool] = None             # 7 · Referral: restituzione al MMG
    referral_note: Optional[str] = None               # 7 · Referral: note / lettera al MMG
    notes: Optional[str] = None                       # Note aggiuntive libere


class WorkupVisit(WorkupVisitBase):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    organization_id: str
    created_by: str
    created_by_name: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.utcnow().isoformat())
    updated_at: Optional[str] = None


@api_router.post("/patients/{patient_id}/workup-visits", response_model=WorkupVisit)
async def create_workup_visit(patient_id: str, payload: WorkupVisitBase, user: dict = Depends(get_current_user)):
    await _verify_patient_in_org(patient_id, user["organization_id"])
    v = WorkupVisit(
        **{**payload.model_dump(), "patient_id": patient_id},
        organization_id=user["organization_id"],
        created_by=user["id"],
        created_by_name=user.get("name"),
    )
    await db.workup_visits.insert_one(v.model_dump())
    return v


@api_router.get("/patients/{patient_id}/workup-visits", response_model=List[WorkupVisit])
async def list_workup_visits(patient_id: str, user: dict = Depends(get_current_user)):
    await _verify_patient_in_org(patient_id, user["organization_id"])
    docs = await db.workup_visits.find(
        {"patient_id": patient_id, "organization_id": user["organization_id"]},
        {"_id": 0},
    ).sort("visit_date", -1).to_list(500)
    return docs


@api_router.put("/workup-visits/{visit_id}", response_model=WorkupVisit)
async def update_workup_visit(visit_id: str, payload: WorkupVisitBase, user: dict = Depends(get_current_user)):
    update_data = {**payload.model_dump(), "updated_at": datetime.utcnow().isoformat()}
    result = await db.workup_visits.update_one(
        {"id": visit_id, "organization_id": user["organization_id"]},
        {"$set": update_data},
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Visita workup non trovata")
    return await db.workup_visits.find_one({"id": visit_id}, {"_id": 0})


@api_router.delete("/workup-visits/{visit_id}")
async def delete_workup_visit(visit_id: str, user: dict = Depends(get_current_user)):
    result = await db.workup_visits.delete_one(
        {"id": visit_id, "organization_id": user["organization_id"]}
    )
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Visita workup non trovata")
    return {"success": True}


@api_router.get("/organization/members")
async def list_org_members(user: dict = Depends(get_current_user)):
    """Return basic info on members of the same organization (for sharing UI)."""
    docs = await db.users.find(
        {"organization_id": user["organization_id"]},
        {"_id": 0, "password_hash": 0},
    ).to_list(500)
    return [
        {"id": d.get("id"), "name": d.get("name"), "email": d.get("email"), "role": d.get("role")}
        for d in docs
    ]


# ==================== PRO TOKENS ====================
@api_router.post("/pro-tokens", response_model=PROToken)
async def create_pro_token(payload: PROTokenCreate, user: dict = Depends(get_current_user)):
    """Doctor creates a short-lived PRO link/QR for a specific patient."""
    await _verify_patient_in_org(payload.patient_id, user["organization_id"])
    invalid = [i for i in payload.instruments if i not in PRO_INSTRUMENTS_ALLOWED]
    if invalid:
        raise HTTPException(status_code=400, detail=f"Strumenti non validi: {invalid}")
    if not payload.instruments:
        raise HTTPException(status_code=400, detail="Devi selezionare almeno uno strumento")
    hours = max(1, min(24 * 60, int(payload.expires_in_hours or 168)))  # max 60 days
    expires = (datetime.now(timezone.utc) + timedelta(hours=hours)).isoformat()
    pt = PROToken(
        patient_id=payload.patient_id,
        organization_id=user["organization_id"],
        instruments=payload.instruments,
        note=payload.note,
        expires_at=expires,
        created_by=user["id"],
        created_by_name=user.get("name"),
    )
    await db.pro_tokens.insert_one(pt.model_dump())
    return pt


@api_router.get("/patients/{patient_id}/pro-tokens", response_model=List[PROToken])
async def list_pro_tokens(patient_id: str, user: dict = Depends(get_current_user)):
    """List all PRO tokens (active+expired+completed) for a patient."""
    await _verify_patient_in_org(patient_id, user["organization_id"])
    docs = await db.pro_tokens.find(
        {"patient_id": patient_id, "organization_id": user["organization_id"]},
        {"_id": 0},
    ).sort("created_at", -1).to_list(500)
    return docs


@api_router.delete("/pro-tokens/{token_id}")
async def delete_pro_token(token_id: str, user: dict = Depends(get_current_user)):
    """Doctor revokes a PRO token (deletes it; pending submissions are lost)."""
    res = await db.pro_tokens.delete_one(
        {"id": token_id, "organization_id": user["organization_id"]}
    )
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Token non trovato")
    return {"success": True}


@api_router.post("/pro-tokens/{token_id}/convert")
async def convert_pro_token(token_id: str, user: dict = Depends(get_current_user)):
    """
    Convert a submitted PRO token into formal assessment(s). Each instrument
    that has a numeric score becomes a single assessment row with index_type
    matching the instrument key.
    """
    pt = await db.pro_tokens.find_one(
        {"id": token_id, "organization_id": user["organization_id"]},
        {"_id": 0},
    )
    if not pt:
        raise HTTPException(status_code=404, detail="Token non trovato")
    if not pt.get("submitted_responses"):
        raise HTTPException(status_code=400, detail="Il paziente non ha ancora compilato il questionario")

    responses = pt["submitted_responses"]
    submission_date = (pt.get("completed_at") or datetime.now(timezone.utc).isoformat())[:10]
    created = []
    for instr, payload in (responses.get("instruments") or {}).items():
        if not isinstance(payload, dict):
            continue
        score = payload.get("score")
        if score is None:
            continue
        doc = {
            "id": str(uuid.uuid4()),
            "patient_id": pt["patient_id"],
            "organization_id": pt["organization_id"],
            "index_type": instr,
            "date": submission_date,
            "score": score,
            "interpretation": payload.get("interpretation"),
            "tender_joints": [],
            "swollen_joints": [],
            "inputs": payload.get("inputs") or {},
            "notes": "Compilato dal paziente via PRO link",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "created_by": user["id"],
            "created_by_name": user.get("name"),
            "source": "patient_reported",
        }
        await db.assessments.insert_one(doc)
        created.append(instr)

    await db.pro_tokens.update_one(
        {"id": token_id}, {"$set": {"converted": True, "converted_at": datetime.now(timezone.utc).isoformat()}}
    )
    return {"success": True, "created": created}


# ===== PUBLIC PRO endpoints (NO AUTH — token-protected) =====
@api_router.get("/public/pro/{token}")
async def public_get_pro(token: str):
    """Patient opens the PRO link → fetch instruments + minimal patient info."""
    pt = await db.pro_tokens.find_one({"token": token}, {"_id": 0})
    if not pt:
        raise HTTPException(status_code=404, detail="Link non valido")
    if pt.get("completed_at"):
        return {
            "status": "already_submitted",
            "completed_at": pt["completed_at"],
            "instruments": pt.get("instruments", []),
        }
    if pt["expires_at"] < datetime.now(timezone.utc).isoformat():
        return {"status": "expired", "expired_at": pt["expires_at"]}

    # Fetch minimal patient context (only first name to confirm to the patient)
    p = await db.patients.find_one({"id": pt["patient_id"]}, {"_id": 0})
    org = await db.organizations.find_one({"id": pt["organization_id"]}, {"_id": 0})
    return {
        "status": "active",
        "patient_first_name": (p or {}).get("nome") or "",
        "patient_code": (p or {}).get("codice_paziente") or "",
        "diagnosis": (p or {}).get("diagnosi") or "",
        "organization_name": (org or {}).get("name") or "",
        "instruments": pt.get("instruments", []),
        "note": pt.get("note"),
        "expires_at": pt["expires_at"],
    }


@api_router.post("/public/pro/{token}/submit")
async def public_submit_pro(token: str, payload: PROSubmit):
    """Patient submits the responses. Idempotent on first submission."""
    pt = await db.pro_tokens.find_one({"token": token}, {"_id": 0})
    if not pt:
        raise HTTPException(status_code=404, detail="Link non valido")
    if pt.get("completed_at"):
        raise HTTPException(status_code=409, detail="Questionario già compilato")
    if pt["expires_at"] < datetime.now(timezone.utc).isoformat():
        raise HTTPException(status_code=410, detail="Link scaduto")

    await db.pro_tokens.update_one(
        {"token": token},
        {
            "$set": {
                "submitted_responses": payload.responses,
                "completed_at": datetime.now(timezone.utc).isoformat(),
            }
        },
    )
    return {"success": True}


# ==================== CONSULT TOKEN ENDPOINTS ====================
@api_router.post("/patients/{patient_id}/consult-token", response_model=ConsultToken)
async def create_consult_token(patient_id: str, payload: ConsultTokenCreate, user: dict = Depends(get_current_user)):
    """Doctor creates a read-only consult link for an external colleague."""
    await _verify_patient_in_org(patient_id, user["organization_id"])
    hours = max(1, min(24 * 365, int(payload.expires_in_hours or 168)))
    expires = (datetime.now(timezone.utc) + timedelta(hours=hours)).isoformat()
    ct = ConsultToken(
        patient_id=patient_id,
        organization_id=user["organization_id"],
        created_by=user["id"],
        created_by_name=user.get("name") or user.get("email"),
        expires_at=expires,
    )
    await db.consult_tokens.insert_one(ct.model_dump())
    return ct


@api_router.get("/patients/{patient_id}/consult-tokens")
async def list_consult_tokens(patient_id: str, user: dict = Depends(get_current_user)):
    await _verify_patient_in_org(patient_id, user["organization_id"])
    docs = await db.consult_tokens.find(
        {"patient_id": patient_id, "organization_id": user["organization_id"]},
        {"_id": 0},
    ).sort("created_at", -1).to_list(50)
    return docs


@api_router.delete("/consult-tokens/{token_id}")
async def delete_consult_token(token_id: str, user: dict = Depends(get_current_user)):
    res = await db.consult_tokens.delete_one(
        {"id": token_id, "organization_id": user["organization_id"]}
    )
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Token non trovato")
    return {"success": True}


@api_router.get("/public/consult/{token}")
async def public_get_consult(token: str):
    """External colleague opens the consult link — returns full read-only patient bundle."""
    ct = await db.consult_tokens.find_one({"token": token}, {"_id": 0})
    if not ct:
        raise HTTPException(status_code=404, detail="Link non valido o scaduto")
    if ct["expires_at"] < datetime.now(timezone.utc).isoformat():
        raise HTTPException(status_code=410, detail="Link scaduto")

    patient_id = ct["patient_id"]
    org_id = ct["organization_id"]

    # Increment view counter asynchronously (best-effort)
    await db.consult_tokens.update_one({"token": token}, {"$inc": {"views": 1}})

    # Fetch all clinical data — pseudonymized (no nome/cognome/CF)
    p = await db.patients.find_one({"id": patient_id}, {"_id": 0}) or {}
    org = await db.organizations.find_one({"id": org_id}, {"_id": 0}) or {}

    patient_safe = {
        "id": p.get("id"),
        "codice_paziente": p.get("codice_paziente"),
        "anno_nascita": p.get("anno_nascita"),
        "sesso": p.get("sesso"),
        "diagnosi": p.get("diagnosi"),
        "diagnosi_secondarie": p.get("diagnosi_secondarie", []),
        "onset_year": p.get("onset_year"),
        "onset_month": p.get("onset_month"),
        "patient_state": p.get("patient_state"),
        "note": p.get("note"),
    }

    assessments = await db.assessments.find(
        {"patient_id": patient_id}, {"_id": 0}
    ).sort("date", -1).to_list(500)

    lab_exams = await db.lab_exams.find(
        {"patient_id": patient_id}, {"_id": 0}
    ).sort("date", -1).to_list(500)

    therapies = await db.therapies.find(
        {"patient_id": patient_id}, {"_id": 0}
    ).sort("start_date", -1).to_list(200)

    criteria_evals = await db.criteria_evaluations.find(
        {"patient_id": patient_id}, {"_id": 0}
    ).sort("date", -1).to_list(100)

    disease_profiles = await db.disease_profiles.find(
        {"patient_id": patient_id}, {"_id": 0}
    ).to_list(20)

    sclero_profiles = await db.sclero_profiles.find(
        {"patient_id": patient_id}, {"_id": 0}
    ).to_list(5)

    workup_visits = await db.workup_visits.find(
        {"patient_id": patient_id}, {"_id": 0}
    ).sort("visit_date", 1).to_list(50)

    return {
        "patient": patient_safe,
        "assessments": assessments,
        "lab_exams": lab_exams,
        "therapies": therapies,
        "criteria_evaluations": criteria_evals,
        "disease_profiles": disease_profiles,
        "sclero_profiles": sclero_profiles,
        "workup_visits": workup_visits,
        "organization_name": org.get("name") or "",
        "expires_at": ct["expires_at"],
        "created_at": ct["created_at"],
    }


# ==================== EXPORT DATABASE ====================
COLLECTIONS_TO_EXPORT = [
    "patients", "assessments", "criteria_evaluations",
    "therapies", "lab_exams", "reminders", "sclero_profiles",
    "disease_profiles",
]


async def _collect_org_data(organization_id: str) -> dict:
    data = {}
    for col in COLLECTIONS_TO_EXPORT:
        docs = await db[col].find({"organization_id": organization_id}, {"_id": 0}).to_list(100000)
        data[col] = docs
    return data


@api_router.get("/export/json")
async def export_json(user: dict = Depends(get_current_user)):
    """Export all org data as a single JSON file."""
    data = await _collect_org_data(user["organization_id"])
    payload = {
        "exported_at": datetime.now(timezone.utc).isoformat(),
        "exported_by": user.get("name"),
        "organization_id": user["organization_id"],
        "version": 1,
        **data,
    }
    body = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
    fname = f"clinimetria_export_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.json"
    return Response(
        content=body,
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


def _flatten_for_csv(value):
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value if value is not None else ""


@api_router.get("/export/csv-zip")
async def export_csv_zip(user: dict = Depends(get_current_user)):
    """Export all org data as a ZIP of CSV files (one per collection)."""
    data = await _collect_org_data(user["organization_id"])
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for col, rows in data.items():
            csv_buf = io.StringIO()
            if rows:
                # Union of all keys across rows for stable schema
                keys = []
                seen = set()
                for r in rows:
                    for k in r.keys():
                        if k not in seen:
                            seen.add(k)
                            keys.append(k)
                writer = csv.DictWriter(csv_buf, fieldnames=keys, extrasaction="ignore")
                writer.writeheader()
                for r in rows:
                    writer.writerow({k: _flatten_for_csv(r.get(k)) for k in keys})
            else:
                csv_buf.write("(no records)\n")
            zf.writestr(f"{col}.csv", csv_buf.getvalue())
        # Manifest
        manifest = {
            "exported_at": datetime.now(timezone.utc).isoformat(),
            "exported_by": user.get("name"),
            "organization_id": user["organization_id"],
            "counts": {col: len(rows) for col, rows in data.items()},
        }
        zf.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))

    buf.seek(0)
    fname = f"clinimetria_export_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.zip"
    return Response(
        content=buf.read(),
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ==================== COHORT EXPORT (XLSX) ====================
def _therapies_active_on(therapies: List[dict], date_str: str) -> List[dict]:
    """Return list of therapies active on a given ISO date (YYYY-MM-DD)."""
    out = []
    for t in therapies:
        start = (t.get("start_date") or "")[:10]
        end = (t.get("end_date") or "")[:10] or None
        if start and start > date_str:
            continue
        if end and end < date_str:
            continue
        # If no start date, skip (can't reason about time)
        if not start:
            continue
        out.append(t)
    return out


def _format_therapy(t: dict) -> str:
    parts = [t.get("drug_name") or "?"]
    if t.get("dose"):
        parts.append(t["dose"])
    if t.get("frequency"):
        parts.append(t["frequency"])
    if t.get("route"):
        parts.append(t["route"])
    return " ".join(parts)


def _days_delta(anchor_iso: Optional[str], visit_iso: Optional[str]):
    """Return integer days delta visit-anchor (negative if before anchor)."""
    if not anchor_iso or not visit_iso:
        return ""
    try:
        a = datetime.fromisoformat(anchor_iso[:10]).date()
        v = datetime.fromisoformat(visit_iso[:10]).date()
        return (v - a).days
    except Exception:
        return ""


async def _load_cohort_data(
    org_id: str, diagnosis: Optional[str]
) -> tuple:
    """
    Load patients (filtered by diagnosis) + their assessments + therapies +
    disease profiles + scleroderma profiles from MongoDB.

    Returns: (patients, assess_by_pid, therap_by_pid, prof_by_pid, sclero_by_pid)
    """
    patient_query: dict = {"organization_id": org_id}
    if diagnosis and diagnosis.strip():
        safe = diagnosis.strip().replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
        patient_query["diagnosi"] = {"$regex": safe, "$options": "i"}
    patients = await db.patients.find(patient_query, {"_id": 0}).to_list(100000)
    patient_ids = [p["id"] for p in patients]

    assess_by_pid: Dict[str, List[dict]] = {}
    therap_by_pid: Dict[str, List[dict]] = {}
    prof_by_pid: Dict[str, Dict[str, dict]] = {}
    sclero_by_pid: Dict[str, dict] = {}

    if patient_ids:
        flt = {"organization_id": org_id, "patient_id": {"$in": patient_ids}}
        for a in await db.assessments.find(flt, {"_id": 0}).to_list(1000000):
            assess_by_pid.setdefault(a["patient_id"], []).append(a)
        for t in await db.therapies.find(flt, {"_id": 0}).to_list(1000000):
            therap_by_pid.setdefault(t["patient_id"], []).append(t)
        for d in await db.disease_profiles.find(flt, {"_id": 0}).to_list(1000000):
            prof_by_pid.setdefault(d["patient_id"], {})[d.get("disease_type", "unknown")] = d
        for s in await db.sclero_profiles.find(flt, {"_id": 0}).to_list(1000000):
            sclero_by_pid[s["patient_id"]] = s

    return patients, assess_by_pid, therap_by_pid, prof_by_pid, sclero_by_pid


def _build_visits_pivot(assess_by_pid: Dict[str, List[dict]]) -> tuple:
    """Group assessments per (patient, date). Returns (visits_by_pid, max_visits, all_indices)."""
    visits_by_pid: Dict[str, List[dict]] = {}
    max_visits = 0
    for pid, alist in assess_by_pid.items():
        by_date: Dict[str, List[dict]] = {}
        for a in alist:
            d = (a.get("date") or a.get("created_at") or "")[:10]
            if d:
                by_date.setdefault(d, []).append(a)
        visits = sorted(({"date": d, "assessments": ass} for d, ass in by_date.items()), key=lambda v: v["date"])
        visits_by_pid[pid] = visits
        if len(visits) > max_visits:
            max_visits = len(visits)

    seen_idx: set = set()
    all_indices: List[str] = []
    for alist in assess_by_pid.values():
        for a in alist:
            it = a.get("index_type")
            if it and it not in seen_idx:
                seen_idx.add(it)
                all_indices.append(it)
    all_indices.sort()
    return visits_by_pid, max_visits, all_indices


def _compute_anchor_dates(
    therap_by_pid: Dict[str, List[dict]], anchor_drug: str
) -> Dict[str, str]:
    """For each patient, return earliest start_date of a therapy whose drug_name contains the anchor (case-insensitive)."""
    anchor_norm = (anchor_drug or "").strip().lower()
    out: Dict[str, str] = {}
    if not anchor_norm:
        return out
    for pid, ther_list in therap_by_pid.items():
        matches = [
            t for t in ther_list
            if t.get("start_date") and anchor_norm in (t.get("drug_name") or "").lower()
        ]
        if matches:
            out[pid] = min(t["start_date"] for t in matches)
    return out


def _collect_profile_columns(
    prof_by_pid: Dict[str, Dict[str, dict]],
    sclero_by_pid: Dict[str, dict],
) -> tuple:
    """Return (profile_cols, sclero_cols): lists of (disease_type|section, key) tuples."""
    profile_cols: List[tuple] = []
    profile_keys_seen: set = set()
    for prof_map in prof_by_pid.values():
        for dtype, doc in prof_map.items():
            for k in (doc.get("data") or {}).keys():
                tag = (dtype, k)
                if tag not in profile_keys_seen:
                    profile_keys_seen.add(tag)
                    profile_cols.append(tag)

    sclero_cols: List[tuple] = []
    sclero_keys_seen: set = set()
    for sdoc in sclero_by_pid.values():
        for sec in ("cutaneous", "antibody", "vascular", "ild", "pah", "gi", "msk"):
            section_data = sdoc.get(sec) or {}
            if not isinstance(section_data, dict):
                continue
            for k in section_data.keys():
                tag = (sec, k)
                if tag not in sclero_keys_seen:
                    sclero_keys_seen.add(tag)
                    sclero_cols.append(tag)
    return profile_cols, sclero_cols


def _build_cohort_columns(
    pseudo: bool,
    profile_cols: List[tuple],
    sclero_cols: List[tuple],
    anchor_norm: bool,
    max_visits: int,
    all_indices: List[str],
) -> List[str]:
    """Build the ordered list of column names for the main 'Coorte' sheet."""
    cols: List[str] = ["codice_paziente"]
    if not pseudo:
        cols.extend(["cognome", "nome", "codice_fiscale", "data_nascita"])
    cols.extend(["anno_nascita", "sesso", "diagnosi", "note", "n_visite"])
    for dtype, k in profile_cols:
        cols.append(f"profilo_{dtype}__{k}")
    for sec, k in sclero_cols:
        cols.append(f"ssc_{sec}__{k}")
    if anchor_norm:
        cols.extend(["anchor_drug", "anchor_t0"])
    for i in range(1, max_visits + 1):
        cols.append(f"t{i}_data")
        if anchor_norm:
            cols.append(f"t{i}_giorni_da_anchor")
        for idx in all_indices:
            cols.append(f"t{i}_{idx}_score")
        cols.append(f"t{i}_terapie_attive")
    return cols


def _build_patient_row(
    p: dict,
    *,
    pseudo: bool,
    visits: List[dict],
    all_ther: List[dict],
    profile_cols: List[tuple],
    sclero_cols: List[tuple],
    prof_map: Dict[str, dict],
    sdoc: Optional[dict],
    all_indices: List[str],
    max_visits: int,
    anchor_drug: Optional[str],
    anchor_t0: Optional[str],
) -> List[Any]:
    """Build the ordered row of cell values for one patient on the 'Coorte' sheet."""
    row: List[Any] = [p.get("codice_paziente") or ""]
    if not pseudo:
        row.extend([
            p.get("cognome") or "",
            p.get("nome") or "",
            p.get("codice_fiscale") or "",
            p.get("data_nascita") or "",
        ])
    row.extend([
        p.get("anno_nascita") or "",
        p.get("sesso") or "",
        p.get("diagnosi") or "",
        (p.get("note") or "").replace("\n", " "),
        len(visits),
    ])
    for dtype, k in profile_cols:
        doc = prof_map.get(dtype)
        val = (doc or {}).get("data", {}).get(k) if doc else None
        row.append(_cell(val))
    for sec, k in sclero_cols:
        val = None
        if sdoc:
            section_data = sdoc.get(sec) or {}
            if isinstance(section_data, dict):
                val = section_data.get(k)
        row.append(_cell(val))
    if anchor_drug:
        row.append(anchor_drug)
        row.append(anchor_t0 or "")
    for i in range(max_visits):
        if i < len(visits):
            v = visits[i]
            row.append(v["date"])
            if anchor_drug:
                row.append(_days_delta(anchor_t0, v["date"]))
            score_map: Dict[str, Any] = {}
            for a in v["assessments"]:
                it = a.get("index_type")
                s = a.get("score")
                if it and s is not None and it not in score_map:
                    score_map[it] = s
            for idx in all_indices:
                row.append(score_map.get(idx, ""))
            active = _therapies_active_on(all_ther, v["date"])
            row.append("; ".join(_format_therapy(t) for t in active) if active else "")
        else:
            row.append("")
            if anchor_drug:
                row.append("")
            for _ in all_indices:
                row.append("")
            row.append("")
    return row


def _style_header_row(ws, ncols: int, row: int = 1) -> None:
    hfont = Font(bold=True, color="FFFFFF", size=10)
    hfill = PatternFill("solid", fgColor="0A2540")
    halign = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for col_idx in range(1, ncols + 1):
        c = ws.cell(row=row, column=col_idx)
        c.font = hfont
        c.fill = hfill
        c.alignment = halign
    ws.row_dimensions[row].height = 30
    ws.freeze_panes = "A2"


def _autosize_columns(ws, cols: List[str]) -> None:
    for col_idx, name in enumerate(cols, start=1):
        max_len = len(name)
        letter = ws.cell(row=1, column=col_idx).column_letter
        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=col_idx).value
            if val is not None:
                L = len(str(val))
                if L > max_len:
                    max_len = L
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 60)


def _write_long_sheet(
    wb,
    *,
    title: str,
    columns: List[str],
    pseudo: bool,
    patients: List[dict],
    items_for_patient,
    sort_key=None,
) -> None:
    """Write a long-format companion sheet (Terapie / Valutazioni)."""
    cols = [c for c in columns if not (pseudo and c in ("cognome", "nome"))]
    ws = wb.create_sheet(title)
    ws.append(cols)
    _style_header_row(ws, len(cols))
    ws.row_dimensions[1].height = 25
    for p in patients:
        items = items_for_patient(p)
        if sort_key:
            items = sorted(items, key=sort_key)
        for it in items:
            row = []
            for k in cols:
                if k == "codice_paziente":
                    row.append(p.get("codice_paziente") or "")
                elif k in ("cognome", "nome", "diagnosi"):
                    row.append(p.get(k) or "")
                else:
                    row.append(_cell(it.get(k)))
            ws.append(row)


@api_router.get("/export/cohort-xlsx")
async def export_cohort_xlsx(
    diagnosis: Optional[str] = None,
    anchor_drug: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    """
    Export an Excel workbook where each row = 1 patient of the cohort
    (optionally filtered by diagnosis, case-insensitive contains), with
    demographics + disease profile + pivoted visits (t1, t2, ...) each
    showing score per index + active therapies at that date.

    If `anchor_drug` is provided (case-insensitive contains match on drug_name),
    each patient's t0 is anchored to the earliest start_date of that drug; an
    extra column `tN_giorni_da_anchor` reports the day delta of each visit
    relative to that anchor (negative = before, positive = after).
    """
    org_id = user["organization_id"]

    patients, assess_by_pid, therap_by_pid, prof_by_pid, sclero_by_pid = await _load_cohort_data(
        org_id, diagnosis
    )
    visits_by_pid, max_visits, all_indices = _build_visits_pivot(assess_by_pid)
    anchor_by_pid = _compute_anchor_dates(therap_by_pid, anchor_drug or "")
    profile_cols, sclero_cols = _collect_profile_columns(prof_by_pid, sclero_by_pid)

    org = await db.organizations.find_one({"id": org_id}, {"_id": 0})
    pseudo = bool((org or {}).get("pseudonymized_mode"))
    anchor_active = bool((anchor_drug or "").strip())

    # Workbook + main sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Coorte"

    cols = _build_cohort_columns(pseudo, profile_cols, sclero_cols, anchor_active, max_visits, all_indices)
    ws.append(cols)
    _style_header_row(ws, len(cols))

    for p in patients:
        ws.append(_build_patient_row(
            p,
            pseudo=pseudo,
            visits=visits_by_pid.get(p["id"], []),
            all_ther=therap_by_pid.get(p["id"], []),
            profile_cols=profile_cols,
            sclero_cols=sclero_cols,
            prof_map=prof_by_pid.get(p["id"], {}),
            sdoc=sclero_by_pid.get(p["id"]),
            all_indices=all_indices,
            max_visits=max_visits,
            anchor_drug=(anchor_drug if anchor_active else None),
            anchor_t0=anchor_by_pid.get(p["id"]),
        ))
    _autosize_columns(ws, cols)

    # Long-format sheets
    _write_long_sheet(
        wb,
        title="Terapie",
        columns=[
            "codice_paziente", "cognome", "nome", "diagnosi", "drug_name", "category",
            "indication", "dose", "frequency", "route", "start_date", "end_date",
            "status", "discontinuation_reason", "auto_discontinued", "notes", "created_by_name",
        ],
        pseudo=pseudo,
        patients=patients,
        items_for_patient=lambda p: therap_by_pid.get(p["id"], []),
    )
    _write_long_sheet(
        wb,
        title="Valutazioni",
        columns=[
            "codice_paziente", "cognome", "nome", "diagnosi", "date",
            "index_type", "score", "interpretation",
            "tender_count", "swollen_count", "created_by_name", "notes",
        ],
        pseudo=pseudo,
        patients=patients,
        items_for_patient=lambda p: assess_by_pid.get(p["id"], []),
        sort_key=lambda x: (x.get("date") or ""),
    )

    # Serialize
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    diag_tag = (diagnosis or "all").strip().replace(" ", "_").replace("/", "_")[:30] or "all"
    fname = f"coorte_{diag_tag}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


def _cell(val):
    """Serialize a cell value safely for openpyxl."""
    if val is None:
        return ""
    if isinstance(val, bool):
        return "sì" if val else "no"
    if isinstance(val, (dict, list)):
        return json.dumps(val, ensure_ascii=False)
    return val


@api_router.get("/export/diagnoses")
async def export_diagnoses(user: dict = Depends(get_current_user)):
    """Return the distinct list of diagnoses present for the organization, to populate the cohort filter."""
    org_id = user["organization_id"]
    vals = await db.patients.distinct("diagnosi", {"organization_id": org_id})
    clean = sorted({(v or "").strip() for v in vals if (v or "").strip()})
    return {"diagnoses": clean}


@api_router.get("/export/drugs")
async def export_drugs(user: dict = Depends(get_current_user)):
    """Return the distinct list of drug names used in therapies, for cohort anchor selection."""
    org_id = user["organization_id"]
    vals = await db.therapies.distinct("drug_name", {"organization_id": org_id})
    clean = sorted({(v or "").strip() for v in vals if (v or "").strip()}, key=lambda s: s.lower())
    return {"drugs": clean}


# ==================== GDPR DE-IDENTIFICATION ENGINE ====================
# Pre-compiled PII patterns for Italian clinical documents.
# Applied to visit text before sending to any AI model (Claude, Gemini, etc.).
# The original text is NEVER stored; only the structured extracted output is persisted.

_RE_CF = re.compile(r'\b[A-Z]{6}\d{2}[A-Z]\d{2}[A-Z]\d{3}[A-Z]\b', re.IGNORECASE)
_RE_EMAIL = re.compile(r'\b[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}\b')
_RE_PHONE = re.compile(
    r'(?<!\d)(\+39\s?|0039\s?)?\(?\d{2,4}\)?\s*[\-\.]?\s*\d{3,4}\s*[\-\.]?\s*\d{3,5}(?!\d)',
    re.IGNORECASE,
)
_RE_BORN = re.compile(
    r'\b(nato|nata|n\.|data\s+di\s+nascita|d\.?o\.?b\.?|nascita)'
    r'[\s/:]+(?:il\s+)?\d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{2,4}\b',
    re.IGNORECASE,
)
_RE_HOSP_HEADER = re.compile(
    r'^(Azienda|A\.?O\.?|Ospedale|IRCCS|ASL|AOU|Policlinico|Università|Universita|Clinica|Istituto|Fondazione|Presidio).+$',
    re.MULTILINE | re.IGNORECASE,
)
_RE_PHYSICIAN_SIG = re.compile(
    r'\b(Dott\.?|Dott\.?ssa|Dr\.?|Prof\.?|Prof\.?ssa)\s+[A-ZÀÈÉÌÒÙ][a-zàèéìòù]+(?:\s+[A-ZÀÈÉÌÒÙ][a-zàèéìòù]+){0,2}\b',
)
_RE_PATIENT_CONTEXT = re.compile(
    r'\b(Paziente|Pz\.?|Sig\.?|Sig\.?ra|Signore?|Signora)[:\s]+[A-ZÀÈÉÌÒÙ][a-zàèéìòù]+(?:\s+[A-ZÀÈÉÌÒÙ][a-zàèéìòù]+){0,2}\b',
)
_RE_ADDRESS = re.compile(
    r'\b(Via|Viale|V\.?le|Corso|C\.?so|Piazza|P\.?za|Vicolo|Contrada|Strada|Str\.)\s+\S+(?:\s+\S+){0,3}[,\s]+\d{5}\b',
    re.IGNORECASE,
)
_RE_HOSP_ID = re.compile(
    r'\b(N\.?\s*ric\.?|N\.?\s*prot\.?|nosologico|NRO|N[°.]\s*paz\.?|ID\s*paz\.?)[:\s]*\d{4,12}\b',
    re.IGNORECASE,
)


def _deidentify_text(
    text: str,
    patient_name: Optional[str] = None,
    patient_surname: Optional[str] = None,
) -> tuple:
    """
    Remove common Italian PII patterns from clinical text before AI processing.
    Returns (cleaned_text, list_of_masked_categories).
    Operates on a copy; the caller's original string is unchanged.
    """
    masked: list = []

    # 1. Patient-specific name/surname from the database record (highest precision)
    for val, label in [(patient_surname, "cognome"), (patient_name, "nome")]:
        if val and len(val) > 2:
            cleaned = re.sub(re.escape(val), "[PAZIENTE]", text, flags=re.IGNORECASE)
            if cleaned != text:
                text = cleaned
                masked.append(label)

    # 2. Codice fiscale (Italian tax ID — 16 chars, deterministic pattern)
    t = _RE_CF.sub("[CF]", text)
    if t != text:
        masked.append("codice_fiscale")
    text = t

    # 3. Email addresses
    t = _RE_EMAIL.sub("[EMAIL]", text)
    if t != text:
        masked.append("email")
    text = t

    # 4. Date of birth in context (nato/nata/data di nascita + date)
    t = _RE_BORN.sub("[DATA_NASCITA]", text)
    if t != text:
        masked.append("data_nascita")
    text = t

    # 5. Hospital/institution identifying headers (full line)
    t = _RE_HOSP_HEADER.sub("[INTESTAZIONE]", text)
    if t != text:
        masked.append("intestazione_ospedaliera")
    text = t

    # 6. Physician signatures (Dott./Dr./Prof. + name)
    t = _RE_PHYSICIAN_SIG.sub("[FIRMA_MEDICO]", text)
    if t != text:
        masked.append("firma_medico")
    text = t

    # 7. Patient name introduced by context keyword (Paziente:, Sig., Pz.)
    t = _RE_PATIENT_CONTEXT.sub("[NOME_PAZIENTE]", text)
    if t != text:
        masked.append("nome_contestuale")
    text = t

    # 8. Italian addresses (Via/Corso/Piazza + street + postal code)
    t = _RE_ADDRESS.sub("[INDIRIZZO]", text)
    if t != text:
        masked.append("indirizzo")
    text = t

    # 9. Hospital record IDs / nosologico numbers
    t = _RE_HOSP_ID.sub("[ID_OSPEDALIERO]", text)
    if t != text:
        masked.append("id_ospedaliero")
    text = t

    return text, masked


# ==================== AI VISIT PARSING ====================
class ParseVisitRequest(BaseModel):
    text: str
    patient_id: Optional[str] = None  # if provided, scope to existing patient


PARSING_SCHEMA_PROMPT = """Sei un assistente clinico esperto in REUMATOLOGIA. Estrai dati strutturati dal testo della visita medica in italiano.
Restituisci ESCLUSIVAMENTE un oggetto JSON valido (no markdown, no commenti, no testo prima/dopo) con questa struttura:

{
  "patient": {
    "nome": "string|null",
    "cognome": "string|null",
    "data_nascita": "YYYY-MM-DD|null",
    "sesso": "M|F|null",
    "codice_fiscale": "string|null",
    "diagnosi": "string|null"
  },
  "assessments": [
    {
      "index_type": "das28_esr|das28_crp|cdai|sdai|basdai|asdas_crp|dapsa|sledai|haq|pasi|basfi|basmi|essdai|esspri|bvas|mmt8|fiqr|mrss|capillaroscopy|schober|lei",
      "date": "YYYY-MM-DD|null",
      "score": "number|null",
      "interpretation": "string|null",
      "tender_count": "number|null",
      "swollen_count": "number|null",
      "inputs": {"any clinically relevant input fields key-value"}
    }
  ],
  "lab_exams": [
    {
      "category": "autoanticorpi|complemento|fase_acuta|emocromo|funzione_organi|urine|altro",
      "date": "YYYY-MM-DD|null",
      "results": [
        {"name": "string", "value": "string", "unit": "string|null", "qualitative": "positivo|negativo|borderline|null", "title": "string|null"}
      ]
    }
  ],
  "therapies": [
    {
      "drug_name": "string",
      "category": "csDMARD|bDMARD|tsDMARD|glucocorticoid|NSAID|analgesic|supportive|other",
      "dose": "string|null",
      "frequency": "string|null",
      "route": "string|null",
      "start_date": "YYYY-MM-DD|null",
      "status": "active|discontinued|completed",
      "notes": "string|null"
    }
  ],
  "sclero_profile": {
    "cutaneous": {"subset": "sine_scleroderma|limited|diffuse|null", "mrss_score": "number|null", "sclerodactyly": "bool|null", "puffy_fingers": "bool|null", "telangiectasias": "bool|null", "calcinosis": "bool|null", "face_involvement": "bool|null"},
    "antibody": {"ana": "string|null", "aca": "neg|pos|borderline|null", "scl70": "neg|pos|borderline|null", "rnap3": "neg|pos|borderline|null", "u3rnp_fibrillarin": "neg|pos|borderline|null", "th_to": "neg|pos|borderline|null", "pm_scl": "neg|pos|borderline|null", "ku": "neg|pos|borderline|null", "u1rnp": "neg|pos|borderline|null"},
    "vascular": {"raynaud": "absent|primary|secondary|null", "raynaud_onset_year": "number|null", "digital_ulcers": "none|past|active_one|active_multiple|null", "capillaroscopy_pattern": "normal|non_specific|early|active|late|null", "pitting_scars": "bool|null", "gangrene": "bool|null", "renal_crisis": "bool|null", "therapy": "string|null"},
    "ild": {"present": "no|yes_stable|yes_progressive|not_assessed|null", "hrct_pattern": "none|nsip|uip|op|mixed|null", "extent": "limited|extensive|null", "fvc_percent": "number|null", "dlco_percent": "number|null", "six_mwt": "number|null", "therapy": "string|null"},
    "pah": {"status": "not_screened|negative|suspected|confirmed|null", "echo_psap": "number|null", "nt_probnp": "number|null", "rhc_done": "yes|no|null", "rhc_mpap": "number|null", "rhc_pcwp": "number|null", "rhc_pvr": "number|null", "who_class": "I|II|III|IV|null", "therapy": "string|null"},
    "gi": {"gerd": "bool|null", "esophageal_dysmotility": "bool|null", "dysphagia": "bool|null", "gavedeformation": "bool|null", "sibo": "bool|null", "intestinal_pseudo_obstruction": "bool|null", "fecal_incontinence": "bool|null", "weight_loss": "bool|null", "therapy": "string|null"},
    "msk": {"arthralgia": "bool|null", "synovitis": "bool|null", "tendon_friction_rubs": "bool|null", "contractures": "bool|null", "myalgia": "bool|null", "myositis": "bool|null", "weakness": "bool|null", "ck_value": "number|null", "therapy": "string|null"}
  },
  "criteria_flags": {
    "haemochromatosis": {
      "iron_fist": "bool|null",
      "joint_onset_before_50": "bool|null",
      "absence_dip_swelling_deformity": "bool|null",
      "mcp_2_5_tenderness": "bool|null",
      "hip_ankle_surgery": "none|hip_only|ankle_only|ankle_and_hip|null",
      "hfe_c282y_homozygous": "bool|null",
      "iron_overload": "bool|null"
    }
  },
  "summary": "breve riepilogo italiano della visita (max 3 righe)"
}

REGOLE:
- Lascia i campi a null se non esplicitamente menzionati nel testo. NON inventare.
- Date in formato ISO YYYY-MM-DD. Se solo l'anno è presente, ometti il campo.
- Per "swollen_count" / "tender_count" estrai numeri di articolazioni se menzionati.
- Per "lei" (Leeds Enthesitis Index): se il testo menziona dolorabilità in entesite/inserzioni
  (es. "Achille destro dolente", "epicondilo laterale sinistro", "condilo femorale mediale dx"),
  popola "inputs.sites" con le chiavi tra: lat_epicondyle_l, lat_epicondyle_r,
  med_femoral_l, med_femoral_r, achilles_l, achilles_r — valore booleano (true se doloroso/positivo).
  Calcola lo score come somma dei siti positivi (range 0–6). Esempio: "Achille bilaterale doloroso,
  epicondilo dx" → sites: {achilles_l: true, achilles_r: true, lat_epicondyle_r: true} score=3.
- Per anticorpi: "pos" se positivo, "neg" se negativo, "borderline" se debole/dubbio.
- Per "sclero_profile" compila SOLO se la diagnosi è sclerosi sistemica/sclerodermia/SSc/VEDOSS.
- Per "criteria_flags.haemochromatosis" compila SOLO se la nota menziona emocromatosi/HFE/sovraccarico di ferro/artropatia da ferro:
  * "iron_fist": true se il paziente NON riesce a chiudere completamente il pugno per limitazione MCP 2-5 (es. "non riesce a chiudere il pugno", "iron fist positivo", "limitazione flessione MCP", "pugno serrato impossibile"). False se la chiusura del pugno è documentata come normale.
  * "joint_onset_before_50": true se l'esordio dei sintomi articolari è prima dei 50 anni.
  * "absence_dip_swelling_deformity": true se le DIP sono descritte SENZA tumefazione e SENZA deformità (cioè non OA-like). False se sono presenti noduli di Heberden/tumefazione DIP.
  * "mcp_2_5_tenderness": true se MCP 2°/3°/4°/5° dolorabili all'esame.
  * "hip_ankle_surgery": "ankle_only" / "hip_only" / "ankle_and_hip" / "none" in base alla storia chirurgica documentata.
  * "hfe_c282y_homozygous": true se la genetica conferma omozigosi C282Y. false se eterozigosi/altro.
  * "iron_overload": true se ferritina elevata, saturazione transferrina elevata, o evidenza biopsia/RMN.
- Output: solo JSON puro. Inizia con { e finisci con }. NIENTE altro."""


async def _call_claude_extract(text: str) -> dict:
    """
    Extract structured clinical data from de-identified visit text using Claude.
    The 'text' argument MUST already be de-identified before calling this function.
    Uses litellm (primary, works with ANTHROPIC_API_KEY).
    Falls back to emergentintegrations if litellm is unavailable.
    Raw text is NEVER logged — only structural metadata (char count, etc.) is allowed in logs.
    """
    import asyncio as _asyncio
    api_key = os.environ.get("ANTHROPIC_API_KEY") or os.environ.get("EMERGENT_LLM_KEY")
    if not api_key:
        raise HTTPException(
            status_code=503,
            detail=(
                "L'analisi AI non è al momento disponibile a causa di un problema di configurazione. "
                "Contattare l'amministratore del sistema."
            ),
        )

    raw: str = ""
    try:
        import litellm as _litellm
        _litellm.drop_params = True
        _litellm.suppress_debug_info = True
        response = await _asyncio.to_thread(
            _litellm.completion,
            model="anthropic/claude-sonnet-4-5-20250929",
            api_key=api_key,
            messages=[
                {"role": "system", "content": PARSING_SCHEMA_PROMPT},
                {"role": "user", "content": f"Testo della visita:\n\n{text}\n\nRestituisci JSON."},
            ],
            max_tokens=8192,
        )
        raw = (response.choices[0].message.content or "").strip()
    except ImportError:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        chat = LlmChat(
            api_key=api_key,
            session_id=f"parse-{uuid.uuid4()}",
            system_message=PARSING_SCHEMA_PROMPT,
        ).with_model("anthropic", "claude-sonnet-4-5-20250929")
        response = await chat.send_message(UserMessage(text=f"Testo della visita:\n\n{text}\n\nRestituisci JSON."))
        raw = (response or "").strip()

    if raw.startswith("```"):
        raw = raw.strip("`")
        if raw.lower().startswith("json"):
            raw = raw[4:].lstrip()
    if "{" in raw and "}" in raw:
        raw = raw[raw.index("{"): raw.rindex("}") + 1]
    try:
        return json.loads(raw)
    except json.JSONDecodeError as e:
        logger.error("Claude JSON parse error: %s", e)
        raise HTTPException(
            status_code=502,
            detail="La risposta del servizio AI non è nel formato atteso. Riprovare.",
        )


@api_router.post("/ai/parse-visit")
async def parse_visit_text(payload: ParseVisitRequest, user: dict = Depends(get_current_user)):
    text = (payload.text or "").strip()
    if len(text) < 30:
        raise HTTPException(status_code=400, detail="Testo troppo breve")
    if len(text) > 25000:
        raise HTTPException(status_code=400, detail="Testo troppo lungo (max 25000 caratteri)")

    # Resolve patient name for high-precision de-identification
    patient_doc = None
    if payload.patient_id:
        patient_doc = await db.patients.find_one(
            {"id": payload.patient_id, "organization_id": user["organization_id"]},
            {"_id": 0, "nome": 1, "cognome": 1},
        )

    # De-identify text BEFORE sending to AI (GDPR Art. 25 — privacy by design).
    # The original text is discarded after this call; only the de-identified copy
    # is forwarded to Claude. Neither version is persisted in the database.
    deidentified_text, masked_categories = _deidentify_text(
        text,
        patient_name=patient_doc.get("nome") if patient_doc else None,
        patient_surname=patient_doc.get("cognome") if patient_doc else None,
    )
    if masked_categories:
        logger.info("AI parse-visit: de-identified %d category/ies: %s", len(masked_categories), masked_categories)

    data = None
    try:
        data = await _call_claude_extract(deidentified_text)
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("AI parse-visit error: %s", type(e).__name__)
        _s = str(e).lower()
        if any(kw in _s for kw in ("api key", "api_key", "invalid", "401", "403", "unauthenticated", "permission")):
            raise HTTPException(
                status_code=503,
                detail=(
                    "L'analisi AI non è al momento disponibile a causa di un problema di configurazione. "
                    "Contattare l'amministratore del sistema."
                ),
            )
        raise HTTPException(
            status_code=502,
            detail="Errore temporaneo del servizio AI. Riprovare tra qualche minuto.",
        )

    return {
        "extracted": data,
        "deidentified": len(masked_categories) > 0,
        "masked_count": len(masked_categories),
    }


# ==================== AI - LAB EXAMS PARSING (PDF/IMAGE) ====================
# Lab schema definition mirrors /app/frontend/src/lib/labPanels.js so that the
# AI returns keys aligned with the existing UI panels and existing data model.
LAB_SCHEMA_KEYS_BY_PANEL = {
    "autoanticorpi": [
        "ana_titolo", "ana_pattern", "anti_dsdna", "anti_sm", "anti_rnp", "anti_ssa_ro", "anti_ssb_la",
        "anti_scl70", "anti_centromero", "anti_rnap3", "anti_pmscl", "anti_ku", "anti_jo1", "anti_mi2",
        "anti_mda5", "anti_tif1g", "anti_nxp2", "anti_srp", "anti_hmgcr",
        "anca_pr3", "anca_mpo", "fr", "acpa_anti_ccp", "anti_mcv",
        "lac", "acl_igg", "acl_igm", "b2gp1_igg", "b2gp1_igm",
    ],
    "complemento": ["c3", "c4", "ch50"],
    "fase_acuta": ["ves", "pcr", "fibrinogeno", "ferritina", "saa"],
    "emocromo": ["wbc", "neutrophils", "lymphocytes", "eosinophils", "monocytes", "hb", "plt"],
    "funzione": ["creatinina", "egfr", "ast", "alt", "ggt", "ldh", "ck", "aldolasi", "urato", "vit_d"],
    "urine": ["proteinuria_24h", "albuminuria", "uacr", "urinary_casts", "hematuria", "sedimento_note"],
}

LAB_PARSING_PROMPT = """Sei un assistente medico esperto in REUMATOLOGIA e LABORATORIO clinico italiano.
RICEVI: il contenuto di un referto di laboratorio (PDF o foto/scansione del referto).
COMPITO: estrarre TUTTI i valori dei test che corrispondono allo schema sottostante e restituire UN SOLO JSON in italiano.

Schema atteso (chiavi standardizzate con cui devi mappare i nomi italiani/inglesi che trovi sul referto):
{
  "date": "YYYY-MM-DD" (se rilevabile la data del prelievo o della firma del referto, altrimenti null),
  "values": {
    "<panel>__<test_key>": {
      "value": numero (oppure null se solo qualitativo),
      "qualitative": "pos|neg|positivo|negativo|debolmente_positivo|null" (solo per test qualitativi tipo Lupus Anticoagulant, ANA pattern, ecc),
      "unit": "stringa, es. mg/L, U/mL, x10^9/L"
    }
  },
  "raw_notes": "note testuali rilevanti dal referto (max 500 caratteri)"
}

Pannelli e chiavi consentite:
- autoanticorpi: ana_titolo, ana_pattern, anti_dsdna, anti_sm, anti_rnp, anti_ssa_ro, anti_ssb_la, anti_scl70, anti_centromero, anti_rnap3, anti_pmscl, anti_ku, anti_jo1, anti_mi2, anti_mda5, anti_tif1g, anti_nxp2, anti_srp, anti_hmgcr, anca_pr3, anca_mpo, fr, acpa_anti_ccp, anti_mcv, lac, acl_igg, acl_igm, b2gp1_igg, b2gp1_igm
- complemento: c3, c4, ch50
- fase_acuta: ves, pcr, fibrinogeno, ferritina, saa
- emocromo: wbc, neutrophils, lymphocytes, eosinophils, monocytes, hb, plt
- funzione: creatinina, egfr, ast, alt, ggt, ldh, ck, aldolasi, urato, vit_d
- urine: proteinuria_24h, albuminuria, uacr, urinary_casts, hematuria, sedimento_note

Regole di mapping italiano/inglese (importanti):
- "VES" / "ESR" → fase_acuta__ves
- "PCR" / "Proteina C reattiva" / "CRP" → fase_acuta__pcr
- "Hb" / "Emoglobina" / "Hemoglobin" → emocromo__hb
- "Globuli bianchi" / "WBC" / "Leucociti" → emocromo__wbc
- "Globuli rossi" / "RBC" → ignorare (non in schema)
- "PLT" / "Piastrine" → emocromo__plt
- "Creatinina" / "Creatinine" → funzione__creatinina
- "GOT" / "AST" → funzione__ast
- "GPT" / "ALT" → funzione__alt
- "γGT" / "GGT" → funzione__ggt
- "CK" / "CPK" / "Creatinchinasi" → funzione__ck
- "C3" / "Complemento C3" → complemento__c3
- "Anti-CCP" / "ACPA" → autoanticorpi__acpa_anti_ccp
- "Fattore Reumatoide" / "FR" / "RF" → autoanticorpi__fr
- "ANA" titolo es 1:160, 1:320 → autoanticorpi__ana_titolo (value=null, "qualitative" può essere il titolo come stringa)
- "Anti-dsDNA" / "Anti-DNA nativo" → autoanticorpi__anti_dsdna
- "Vitamina D" / "25(OH)D" → funzione__vit_d
- "Acido urico" / "Uricemia" → funzione__urato
- "Ferritina" → fase_acuta__ferritina
- "eGFR" / "Velocità filtrazione glomerulare" / "GFR" → funzione__egfr
- "Proteinuria 24h" / "Proteinuria delle 24 ore" → urine__proteinuria_24h

Linee guida operative:
- INSERISCI nel JSON SOLO i test effettivamente presenti nel referto. Tutti gli altri li ometti (NON null fittizi).
- Per test qualitativi (Lupus Anticoagulant, ANA pattern, anti-Sm, ecc), valorizza "qualitative" e lascia "value" a null.
- Conserva l'unità di misura come scritta sul referto.
- Se il valore è "<5" o ">300" tieni il numero principale e mettilo in "value", aggiungi un commento in raw_notes.
- Se il referto contiene SOLO immagini (foto del referto) leggi via OCR e applica le stesse regole.
- NON inventare valori non presenti nel referto.
- IGNORA e NON includere in "raw_notes" dati identificativi del paziente (nome, cognome, codice fiscale, data di nascita, indirizzo, telefono). Riporta solo informazioni cliniche (valori, unità, note metodologiche).
- Output: SOLO JSON puro. Inizia con { e finisci con }. Niente altro."""


def _extract_pdf_text_local(pdf_bytes: bytes) -> str:
    """
    Extract plain text from a PDF entirely on the server using pypdf — no network call,
    no external service. Works for digitally-generated PDFs (the vast majority of
    Italian lab reports). Returns empty string for scanned image-only PDFs.

    GDPR note: extracted text is used only transiently for de-identification and AI
    parsing. It is NEVER written to disk, logged, or persisted in any form.
    """
    try:
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(pdf_bytes))
        pages: list = []
        for page in reader.pages:
            try:
                t = page.extract_text() or ""
                if t.strip():
                    pages.append(t)
            except Exception:
                pass
        return "\n".join(pages)
    except Exception as exc:
        logger.warning("PDF local text extraction failed (%s) — will treat as image.", type(exc).__name__)
        return ""


async def _gemini_extract_lab_text(deidentified_text: str) -> dict:
    """
    Extract structured lab values from already-de-identified text via Gemini Flash models.

    Model fallback chain (cheapest / free-tier first — no Pro required):
        1. gemini-2.0-flash          — primary; generous free tier (1 500 RPD)
        2. gemini-1.5-flash          — first fallback
        3. gemini-1.5-flash-latest   — second fallback (alias kept up-to-date by Google)

    Uses litellm REST path (not gRPC) for better compatibility with API-Studio keys.

    GDPR note: 'deidentified_text' must already have been processed by _deidentify_text().
    This function NEVER logs text content — only structural metadata (char count, model used).
    """
    import asyncio as _asyncio
    import litellm as _litellm

    api_key = (
        os.environ.get("GEMINI_API_KEY")
        or os.environ.get("GOOGLE_API_KEY")
        or os.environ.get("EMERGENT_LLM_KEY")
    )
    if not api_key:
        raise HTTPException(
            status_code=503,
            detail=(
                "L'estrazione AI non è al momento disponibile a causa di un problema di configurazione. "
                "Contattare l'amministratore del sistema."
            ),
        )

    _litellm.drop_params = True
    _litellm.suppress_debug_info = True

    # Ordered from cheapest/most-available to slightly more capable.
    # None of these require billing beyond the free tier.
    FLASH_MODELS = [
        "gemini/gemini-2.0-flash",
        "gemini/gemini-1.5-flash",
        "gemini/gemini-1.5-flash-latest",
    ]

    messages = [
        {"role": "system", "content": LAB_PARSING_PROMPT},
        {
            "role": "user",
            "content": (
                "Estrai i valori di laboratorio dal seguente testo di referto "
                "(già de-identificato lato server — nessun dato paziente presente). "
                "Segui lo schema nel system prompt. Restituisci SOLO JSON valido.\n\n"
                + deidentified_text
            ),
        },
    ]

    last_exc: Exception | None = None
    for model_name in FLASH_MODELS:
        try:
            response = await _asyncio.to_thread(
                _litellm.completion,
                model=model_name,
                api_key=api_key,
                messages=messages,
                max_tokens=4096,
            )
            raw = (response.choices[0].message.content or "").strip()
            logger.info("Lab AI text-parse succeeded with model %s (%d chars out)", model_name, len(raw))
            break  # success — exit the fallback loop
        except Exception as exc:
            _s = str(exc).lower()
            logger.warning("Lab AI: model %s failed (%s)", model_name, type(exc).__name__)

            # Auth / key errors — no point trying other models
            if any(kw in _s for kw in ("api key", "api_key", "invalid", "401", "403",
                                        "unauthenticated", "permission", "key not valid")):
                logger.error("Lab AI: auth error on model %s", model_name)
                raise HTTPException(
                    status_code=503,
                    detail=(
                        "L'estrazione AI non è al momento disponibile a causa di un problema di configurazione. "
                        "Contattare l'amministratore del sistema."
                    ),
                )

            # Quota / rate errors — try next model in chain
            if any(kw in _s for kw in ("quota", "429", "resource_exhausted",
                                        "resourceexhausted", "rate", "limit")):
                last_exc = exc
                continue  # try next model

            # Unexpected error — still try next model but log as error
            logger.error("Lab AI: unexpected error on model %s: %s", model_name, type(exc).__name__)
            last_exc = exc
            continue
    else:
        # All models exhausted
        _s = str(last_exc).lower() if last_exc else ""
        if any(kw in _s for kw in ("quota", "429", "resource_exhausted", "limit")):
            raise HTTPException(
                status_code=429,
                detail=(
                    "Quota AI esaurita su tutti i modelli disponibili. "
                    "Riprovare più tardi o abilitare la fatturazione nel progetto Google Cloud."
                ),
            )
        raise HTTPException(
            status_code=502,
            detail="Errore del servizio AI. Riprovare tra qualche minuto.",
        )

    # Parse the JSON response
    if raw.startswith("```"):
        raw = raw.strip("`")
        if raw.lower().startswith("json"):
            raw = raw[4:].lstrip()
    if "{" in raw and "}" in raw:
        raw = raw[raw.index("{"):raw.rindex("}") + 1]
    try:
        return json.loads(raw)
    except json.JSONDecodeError as exc:
        logger.error("Lab AI JSON parse error: %s", exc)
        raise HTTPException(
            status_code=502,
            detail="La risposta del servizio AI non è nel formato atteso. Riprovare.",
        )


def _resolve_lab_file_mime(content_type: str, fname: str) -> tuple:
    """Return (mime, ext) tuple for supported lab file types, or raise 400."""
    ct = (content_type or "").lower()
    fn = fname.lower()
    if ct == "application/pdf" or fn.endswith(".pdf"):
        return "application/pdf", ".pdf"
    if ct in ("image/jpeg", "image/jpg") or fn.endswith((".jpg", ".jpeg")):
        return "image/jpeg", ".jpg"
    if ct == "image/png" or fn.endswith(".png"):
        return "image/png", ".png"
    if ct == "image/webp" or fn.endswith(".webp"):
        return "image/webp", ".webp"
    raise HTTPException(
        status_code=400,
        detail="Formato non supportato. Carica PDF, JPEG, PNG o WEBP.",
    )


async def _gemini_extract_lab(tmp_path: str, mime: str) -> dict:
    """Send the lab file to Gemini 2.5 Pro and return the parsed JSON dict."""
    from emergentintegrations.llm.chat import LlmChat, UserMessage, FileContentWithMimeType

    api_key = os.environ.get("EMERGENT_LLM_KEY")
    if not api_key:
        raise HTTPException(
            status_code=503,
            detail=(
                "L'estrazione AI non è al momento disponibile a causa di un problema di configurazione. "
                "Contattare l'amministratore del sistema."
            ),
        )

    chat = LlmChat(
        api_key=api_key,
        session_id=f"lab-{uuid.uuid4()}",
        system_message=LAB_PARSING_PROMPT,
    ).with_model("gemini", "gemini-2.5-pro")

    attachment = FileContentWithMimeType(file_path=tmp_path, mime_type=mime)
    user_message = UserMessage(
        text=(
            "Estrai i valori di laboratorio da questo referto secondo lo schema "
            "indicato nel system prompt. Restituisci SOLO JSON valido."
        ),
        file_contents=[attachment],
    )

    try:
        response = await chat.send_message(user_message)
    except Exception as e:
        logger.exception("Lab AI multimodal parse error: %s", type(e).__name__)
        _s = str(e).lower()
        if any(kw in _s for kw in ("api key", "api_key", "invalid", "401", "403", "unauthenticated", "permission")):
            raise HTTPException(
                status_code=503,
                detail=(
                    "L'estrazione AI non è al momento disponibile a causa di un problema di configurazione. "
                    "Contattare l'amministratore del sistema."
                ),
            )
        raise HTTPException(
            status_code=502,
            detail="Errore temporaneo del servizio AI. Riprovare tra qualche minuto.",
        )

    raw = (response or "").strip()
    if raw.startswith("```"):
        raw = raw.strip("`")
        if raw.lower().startswith("json"):
            raw = raw[4:].lstrip()
    if "{" in raw and "}" in raw:
        raw = raw[raw.index("{") : raw.rindex("}") + 1]
    try:
        return json.loads(raw)
    except json.JSONDecodeError as e:
        raise HTTPException(status_code=502, detail=f"Risposta AI non valida: {e}")


def _reorganize_lab_values(values: dict) -> tuple:
    """Reorganize "panel__test" flat keys into nested {panel: {test_key: val}}.

    Returns (reorganized_dict, unmatched_keys_list).
    """
    reorganized: Dict[str, Dict[str, Any]] = {}
    unmatched: List[str] = []
    for key, val in (values or {}).items():
        if "__" not in key:
            unmatched.append(key)
            continue
        panel, test_key = key.split("__", 1)
        if panel not in LAB_SCHEMA_KEYS_BY_PANEL:
            unmatched.append(key)
            continue
        if test_key not in LAB_SCHEMA_KEYS_BY_PANEL[panel]:
            unmatched.append(key)
            continue
        reorganized.setdefault(panel, {})[test_key] = val
    return reorganized, unmatched


@api_router.post("/ai/parse-lab")
async def parse_lab_file(
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user),
):
    """
    Local-first, GDPR-compliant lab report parsing pipeline (GDPR Art. 25).

    Priority order:
        1. Local text extraction (pdftotext → pypdf for PDFs; Tesseract OCR for images)
           + deterministic rule-based regex parser.  Free, no API key, no network.
        2. AI fallback (Gemini Flash) — ONLY when BOTH:
               a) An API key is set, AND
               b) Local parse returned < 2 values or all items are low-confidence.
           AI receives ONLY de-identified text — never raw document bytes.

    GDPR guarantees:
        - File bytes processed in-memory only; never written to permanent storage.
        - De-identification applied before any external call.
        - Text discarded immediately after parsing.
        - Nothing PII-related is ever logged.
    """
    import asyncio as _asyncio
    from lab_parser import (
        extract_text_from_pdf, extract_text_from_image,
        parse_lab_text, items_to_panels, needs_ai_fallback,
    )

    if not file.filename:
        raise HTTPException(status_code=400, detail="File non valido")

    mime, ext = _resolve_lab_file_mime(file.content_type or "", file.filename)

    contents = await file.read()
    if not contents:
        raise HTTPException(status_code=400, detail="File vuoto")
    if len(contents) > 15 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File troppo grande (max 15 MB)")

    # ── Step 1: Extract raw text locally ─────────────────────────────────────
    raw_text: str = ""
    text_method: str = "none"

    if mime == "application/pdf":
        raw_text, text_method = await _asyncio.to_thread(extract_text_from_pdf, contents)
    else:
        raw_text, text_method = await _asyncio.to_thread(extract_text_from_image, contents, mime)
        # Images: strict-privacy check if OCR found nothing
        if not raw_text.strip():
            org = await db.organizations.find_one(
                {"id": user["organization_id"]}, {"_id": 0, "ai_strict_privacy": 1}
            )
            if org and org.get("ai_strict_privacy"):
                del contents
                raise HTTPException(
                    status_code=403,
                    detail=(
                        "Modalità AI Strict Privacy attiva: il testo non è stato estratto localmente "
                        "dall'immagine (Tesseract non disponibile) e l'invio all'AI è bloccato. "
                        "Carica un referto PDF digitale per la piena conformità GDPR."
                    ),
                )

    del contents  # release raw bytes immediately — never referenced again

    logger.info(
        "Lab parse step-1: mime=%s text_method=%s text_chars=%d",
        mime, text_method, len(raw_text),
    )

    # ── Step 2: Rule-based local parse ────────────────────────────────────────
    local_parse = await _asyncio.to_thread(parse_lab_text, raw_text) if raw_text.strip() else {"date": None, "items": [], "raw_notes": ""}
    local_panels, local_conf_map, _ = items_to_panels(local_parse["items"])

    local_result = {
        "panels":           local_panels,
        "confidence_map":   local_conf_map,
        "date":             local_parse["date"],
        "raw_notes":        local_parse["raw_notes"],
        "items":            local_parse["items"],
    }

    logger.info(
        "Lab parse step-2 (rules): items=%d panels=%s",
        len(local_parse["items"]), list(local_panels.keys()),
    )

    # ── Step 3: Optional AI fallback ─────────────────────────────────────────
    has_api_key = bool(
        os.environ.get("GEMINI_API_KEY")
        or os.environ.get("GOOGLE_API_KEY")
        or os.environ.get("EMERGENT_LLM_KEY")
    )

    final_panels      = local_panels
    final_conf_map    = local_conf_map
    final_date        = local_parse["date"]
    final_notes       = local_parse["raw_notes"]
    extraction_method = f"local/{text_method}"

    if has_api_key and needs_ai_fallback(local_result) and raw_text.strip():
        logger.info(
            "Lab parse step-3 (AI fallback): local items=%d — calling Gemini Flash",
            len(local_parse["items"]),
        )
        try:
            # De-identify before sending to AI — no PII transmitted
            deidentified_text, masked = _deidentify_text(raw_text)
            logger.info(
                "Lab parse AI: %d de-id categories applied, %d chars → AI",
                len(masked), len(deidentified_text),
            )
            ai_raw = await _gemini_extract_lab_text(deidentified_text)
            del deidentified_text  # discard immediately

            ai_reorganized, _ = _reorganize_lab_values(ai_raw.get("values") or {})

            # Merge: AI fills in panels the local parser missed
            for panel_key, ai_vals in ai_reorganized.items():
                if panel_key not in final_panels:
                    final_panels[panel_key] = ai_vals
                    for test_key in ai_vals:
                        final_conf_map[f"{panel_key}__{test_key}"] = "medium"
                else:
                    # Overlay individual tests the local parser didn't catch
                    for test_key, ai_val in ai_vals.items():
                        if test_key not in final_panels[panel_key]:
                            final_panels[panel_key][test_key] = ai_val
                            final_conf_map[f"{panel_key}__{test_key}"] = "medium"

            if not final_date:
                final_date = ai_raw.get("date")
            if not final_notes:
                final_notes = ai_raw.get("raw_notes", "")
            extraction_method = f"local+ai/{text_method}"
            logger.info("Lab parse step-3 (AI): merged %d panels", len(ai_reorganized))

        except HTTPException:
            logger.warning("Lab parse AI fallback: API error — using local result only")
        except Exception as exc:
            logger.warning("Lab parse AI fallback: %s — using local result only", type(exc).__name__)

    del raw_text  # discard extracted text

    return {
        "date":              final_date,
        "panels":            final_panels,
        "confidence_map":    final_conf_map,
        "raw_notes":         final_notes,
        "unmatched_keys":    [],
        "filename":          file.filename,
        "extraction_method": extraction_method,
        "local_item_count":  len(local_parse["items"]),
    }


# ==================== STATS ====================
@api_router.get("/stats")
async def stats(user: dict = Depends(get_current_user)):
    org_id = user["organization_id"]
    patients_count = await db.patients.count_documents({"organization_id": org_id})
    assessments_count = await db.assessments.count_documents({"organization_id": org_id})
    recent = await db.assessments.find({"organization_id": org_id}, {"_id": 0}).sort("created_at", -1).limit(5).to_list(5)
    return {"patients": patients_count, "assessments": assessments_count, "recent_assessments": recent}


# ==================== SCLERODERMA PROFILE ENDPOINTS ====================
@api_router.get("/patients/{patient_id}/sclero-profile")
async def get_sclero_profile(patient_id: str, user: dict = Depends(get_current_user)):
    await _verify_patient_in_org(patient_id, user["organization_id"])
    doc = await db.sclero_profiles.find_one(
        {"patient_id": patient_id, "organization_id": user["organization_id"]},
        {"_id": 0},
    )
    return doc  # null if not exists


@api_router.put("/patients/{patient_id}/sclero-profile", response_model=ScleroProfile)
async def upsert_sclero_profile(
    patient_id: str,
    payload: ScleroProfileBase,
    user: dict = Depends(get_current_user),
):
    await _verify_patient_in_org(patient_id, user["organization_id"])
    if payload.patient_id != patient_id:
        raise HTTPException(status_code=400, detail="patient_id mismatch")

    existing = await db.sclero_profiles.find_one(
        {"patient_id": patient_id, "organization_id": user["organization_id"]},
        {"_id": 0},
    )
    now = datetime.now(timezone.utc).isoformat()
    if existing:
        update_data = payload.model_dump(exclude={"patient_id"})
        update_data["updated_at"] = now
        update_data["updated_by_name"] = user.get("name")
        await db.sclero_profiles.update_one(
            {"id": existing["id"]},
            {"$set": update_data},
        )
        return await db.sclero_profiles.find_one({"id": existing["id"]}, {"_id": 0})

    profile = ScleroProfile(
        **payload.model_dump(),
        organization_id=user["organization_id"],
        created_by=user["id"],
        created_by_name=user.get("name"),
        updated_by_name=user.get("name"),
    )
    await db.sclero_profiles.insert_one(profile.model_dump())
    return profile


# ==================== GENERIC DISEASE PROFILE (RA / SpA / ...) ====================
ALLOWED_DISEASE_TYPES = {"ra", "spa", "sle", "aav", "sjogren", "myositis", "pmr_lvv", "comorbidities", "prima_visita", "clinical_cockpit"}


class DiseaseProfileBase(BaseModel):
    patient_id: str
    data: Dict[str, Any] = Field(default_factory=dict)


class DiseaseProfile(DiseaseProfileBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    organization_id: str
    disease_type: str
    created_by: str
    created_by_name: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    updated_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    updated_by_name: Optional[str] = None


def _check_disease_type(t: str):
    if t not in ALLOWED_DISEASE_TYPES:
        raise HTTPException(status_code=400, detail=f"Disease type '{t}' non supportato")


@api_router.get("/patients/{patient_id}/disease-profile/{disease_type}")
async def get_disease_profile(
    patient_id: str, disease_type: str, user: dict = Depends(get_current_user)
):
    _check_disease_type(disease_type)
    await _verify_patient_in_org(patient_id, user["organization_id"])
    doc = await db.disease_profiles.find_one(
        {
            "patient_id": patient_id,
            "organization_id": user["organization_id"],
            "disease_type": disease_type,
        },
        {"_id": 0},
    )
    return doc  # null if not exists


@api_router.delete("/patients/{patient_id}/disease-profile/{disease_type}", status_code=204)
async def delete_disease_profile(
    patient_id: str,
    disease_type: str,
    user: dict = Depends(get_current_user),
):
    _check_disease_type(disease_type)
    await _verify_patient_in_org(patient_id, user["organization_id"])
    await db.disease_profiles.delete_one({
        "patient_id": patient_id,
        "organization_id": user["organization_id"],
        "disease_type": disease_type,
    })
    return None


@api_router.put("/patients/{patient_id}/disease-profile/{disease_type}", response_model=DiseaseProfile)
async def upsert_disease_profile(
    patient_id: str,
    disease_type: str,
    payload: DiseaseProfileBase,
    user: dict = Depends(get_current_user),
):
    _check_disease_type(disease_type)
    await _verify_patient_in_org(patient_id, user["organization_id"])
    if payload.patient_id != patient_id:
        raise HTTPException(status_code=400, detail="patient_id mismatch")

    existing = await db.disease_profiles.find_one(
        {
            "patient_id": patient_id,
            "organization_id": user["organization_id"],
            "disease_type": disease_type,
        },
        {"_id": 0},
    )
    now = datetime.now(timezone.utc).isoformat()
    if existing:
        await db.disease_profiles.update_one(
            {"id": existing["id"]},
            {
                "$set": {
                    "data": payload.data,
                    "updated_at": now,
                    "updated_by_name": user.get("name"),
                }
            },
        )
        return await db.disease_profiles.find_one({"id": existing["id"]}, {"_id": 0})

    profile = DiseaseProfile(
        **payload.model_dump(),
        disease_type=disease_type,
        organization_id=user["organization_id"],
        created_by=user["id"],
        created_by_name=user.get("name"),
        updated_by_name=user.get("name"),
    )
    await db.disease_profiles.insert_one(profile.model_dump())
    return profile


app.include_router(api_router)


# ==================== STARTUP ====================
@app.on_event("startup")
async def startup_event():
    # Indexes
    try:
        await db.users.create_index("email", unique=True)
        await db.users.create_index("id", unique=True)
        await db.organizations.create_index("invite_code", unique=True)
        await db.organizations.create_index("id", unique=True)
        await db.patients.create_index("organization_id")
        await db.assessments.create_index([("patient_id", 1), ("date", -1)])
        await db.criteria_evaluations.create_index([("patient_id", 1), ("date", -1)])
        await db.therapies.create_index([("patient_id", 1)])
        await db.sclero_profiles.create_index([("patient_id", 1)], unique=True)
        await db.disease_profiles.create_index(
            [("patient_id", 1), ("disease_type", 1)], unique=True
        )
        await db.workup_visits.create_index([("patient_id", 1), ("visit_date", -1)])
    except Exception as e:
        logger.warning(f"Index creation: {e}")

    # Default organization
    default_org_name = os.environ.get("DEFAULT_ORG_NAME", "UO Reumatologia")
    default_org = await db.organizations.find_one({"name": default_org_name}, {"_id": 0})
    if not default_org:
        new_org = Organization(name=default_org_name)
        await db.organizations.insert_one(new_org.model_dump())
        default_org = new_org.model_dump()
    default_org_id = default_org["id"]

    # Migrate existing patients/assessments without organization_id
    await db.patients.update_many({"organization_id": {"$exists": False}}, {"$set": {"organization_id": default_org_id, "created_by": "system"}})
    await db.assessments.update_many({"organization_id": {"$exists": False}}, {"$set": {"organization_id": default_org_id, "created_by": "system"}})
    await db.criteria_evaluations.update_many({"organization_id": {"$exists": False}}, {"$set": {"organization_id": default_org_id, "created_by": "system"}})
    await db.therapies.update_many({"organization_id": {"$exists": False}}, {"$set": {"organization_id": default_org_id, "created_by": "system"}})

    # Seed admin
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@clinimetria.it")
    admin_password = os.environ.get("ADMIN_PASSWORD", "admin123")
    existing_admin = await db.users.find_one({"email": admin_email})
    if not existing_admin:
        await db.users.insert_one({
            "id": str(uuid.uuid4()),
            "email": admin_email,
            "name": "Admin",
            "password_hash": hash_password(admin_password),
            "role": "admin",
            "organization_id": default_org_id,
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
    elif not verify_password(admin_password, existing_admin["password_hash"]):
        await db.users.update_one(
            {"email": admin_email},
            {"$set": {"password_hash": hash_password(admin_password)}},
        )

    # Write test credentials
    try:
        memory_dir = Path("/app/memory")
        memory_dir.mkdir(parents=True, exist_ok=True)
        org = await db.organizations.find_one({"id": default_org_id}, {"_id": 0})
        creds_path = memory_dir / "test_credentials.md"
        creds_path.write_text(
            f"# Test Credentials\n\n"
            f"## Admin (Default UO)\n"
            f"- Email: `{admin_email}`\n"
            f"- Password: `{admin_password}`\n"
            f"- Role: admin\n"
            f"- Organization: `{org['name']}`\n"
            f"- Invite code: `{org['invite_code']}`\n\n"
            f"## Auth endpoints\n"
            f"- POST /api/auth/register\n"
            f"- POST /api/auth/login\n"
            f"- POST /api/auth/logout\n"
            f"- GET /api/auth/me\n"
            f"- POST /api/auth/refresh\n"
        )
    except Exception as e:
        logger.warning(f"Failed to write test_credentials.md: {e}")


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()


# ==================== SPA FALLBACK ====================
# Production:  serve the compiled React build from frontend/build/
# Development: transparently proxy to the React dev server on port 5000
#              (port 8000/FastAPI is mapped to external port 80 in Replit dev)

_FRONTEND_BUILD = ROOT_DIR.parent / "frontend" / "build"
_SPA_DEV_SERVER = "http://localhost:5000"
_HOP_BY_HOP = {"connection", "transfer-encoding", "te", "upgrade",
                "proxy-authorization", "proxy-authenticate", "trailers"}


@app.api_route("/{path:path}", methods=["GET", "HEAD"], include_in_schema=False)
async def spa_fallback(path: str, request: Request):
    # ── PRODUCTION: serve compiled static files ──────────────────────────────
    if (_FRONTEND_BUILD / "index.html").exists():
        # Try to serve an exact file from the build directory (JS, CSS, images…)
        if path:
            candidate = (_FRONTEND_BUILD / path).resolve()
            build_root = _FRONTEND_BUILD.resolve()
            try:
                candidate.relative_to(build_root)   # guard path traversal
                if candidate.is_file():
                    return FileResponse(str(candidate))
            except ValueError:
                pass
        # Everything else (React routes) → index.html (no-cache so browsers always re-fetch)
        return FileResponse(
            str(_FRONTEND_BUILD / "index.html"),
            headers={"Cache-Control": "no-cache, no-store, must-revalidate", "Pragma": "no-cache"},
        )

    # ── DEVELOPMENT: proxy to React dev server ───────────────────────────────
    target = f"{_SPA_DEV_SERVER}/{path}"
    query = request.url.query
    if query:
        target += f"?{query}"
    # Force Accept-Encoding: identity so the dev server never compresses.
    # This prevents the httpx auto-decompress/Content-Encoding mismatch that
    # would otherwise produce garbled binary output in the browser.
    _strip_req = _HOP_BY_HOP | {"host", "accept-encoding"}
    forward_headers = {
        k: v for k, v in request.headers.items()
        if k.lower() not in _strip_req
    }
    forward_headers["accept-encoding"] = "identity"
    try:
        async with httpx.AsyncClient(timeout=30.0) as hc:
            resp = await hc.request(
                request.method, target,
                headers=forward_headers, follow_redirects=True,
            )
        _strip_resp = _HOP_BY_HOP | {"content-encoding", "content-length"}
        resp_headers = {
            k: v for k, v in resp.headers.items()
            if k.lower() not in _strip_resp
        }
        return Response(
            content=resp.content,
            status_code=resp.status_code,
            headers=resp_headers,
            media_type=resp.headers.get("content-type"),
        )
    except Exception as exc:
        logger.error("[spa_fallback] %s → %s: %s", request.url.path, target, exc)
        return Response(
            content=(
                "<html><body><b>Dev server unavailable.</b><br>"
                "Make sure the React dev server is running on port 5000.<br>"
                f"<code>{exc}</code></body></html>"
            ),
            status_code=502,
            media_type="text/html",
        )


# ==================== CORS ====================
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)
